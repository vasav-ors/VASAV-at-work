"""
===================================================================================================
PILE DRIVEABILITY ANALYSIS - INTERACTIVE PLOTTING TOOL
===================================================================================================

PURPOSE:
--------
This script extracts, processes, and visualizes pile driving analysis results for offshore wind
turbine monopile installations. It creates comprehensive interactive plots showing Static Resistance
to Driving (SRD), blowcount rates, energy requirements, and driveability assessment metrics.

EXECUTION ORDER:
----------------
1. USER CONFIGURATION
   - Set threshold values (hard driving, refusal blowcount rates)
   - Define file paths (monopile weights, root directory, output directory)
   - Configure gripper penetration and refusal risk data sources

2. DATA COLLECTION PHASE (main function)
   a) Scan position folders (A01, A02, etc.) in the monopile root directory
   b) Prompt user to select positions to analyze (or 'all')
   c) Load gripper penetration data for selected positions only (optimization)
   d) Parse driveability CSV files for selected positions:
      - Extract multiple tables from each results_PileDrivingAnalysis-{Position}.csv
      - Identify available SRD methods (MD, MY, AH) and soil bounds (lb, be, ub)
      - Extract target penetration depths for refusal risk calculations
   e) Prompt user to select SRD methods and soil bounds to plot
   f) Load refusal risk assessment data from Excel file (1hr, 24hr, 48hr, 7days pause durations)

3. POSITION PROCESSING LOOP
   For each selected position:
   a) Extract soil profile data from input_Position_{Position}.csv file:
      - Read soil layers with qc values, colors, and geoUnit classifications
      - Calculate depth below seabed using zSeabed reference
      - Prepare step plot data for qc profile overlay

   b) Call plot_driveability_results() which:
      - Creates 2x4 subplot grid (3 plots per row + 1 table per row)
      - Row 1: SRD [MN], Blowcount Rate, Input Energy, Info Table
      - Row 2: Total Energy [GJ], Cumulative Blows, SRD [kN] (log scale), Assessment Table
      - Plots all selected method/bound combinations with color-coded traces
      - Overlays soil profile (qc, layers, geoUnits) on SRD subplot
      - Adds target depth line, hard driving thresholds, refusal criteria
      - Adds horizontal lines for gripper release and MP abandonment depths
      - Adds horizontal lines for refusal risk depths (1hr, 24hr, 48hr, 7days)

   c) Calculate self-weight penetration and pile run assessments:
      - SWP MP + ILT (Internal Lifting Tool): Depth where SRD = MP + ILT weight
      - Pile run at hammer placement: Check if SRD < threshold at hammer placement
      - SWP MP + Hammer: Depth where SRD = MP + Hammer weight
      - Pile run risk top: First depth where SRD drops below hammer weight (risk initiation)
      - Pile run risk bottom: Depth where SRD recovers above threshold (risk zone end)
      - CONSERVATIVE LOGIC: When multiple methods selected, use most conservative value
        * SWP depths: DEEPEST (pile penetrates furthest)
        * Pile run risks: SHALLOWEST for top (earliest risk), DEEPEST for bottom (longest zone)

   d) Populate information table with:
      - Position name, monopile weight, target depth, hammer details
      - Minimum penetrations for gripper release and MP abandonment
      - Refusal risk depths for different installation pause durations

   e) Save plot as: Installation_Driveability_{Position}.html
   f) Display interactive plot in browser

KEY FEATURES:
-------------
- Multi-method comparison: Compare MD (MonoDrive), MY (Maynard), AH (Alm & Hamre)
- Multi-bound analysis: Lower bound (lb), Best estimate (be), Upper bound (ub)
- Synchronized y-axes: All subplots zoom together for easy comparison
- Soil profile overlay: qc profile with color-coded layers and geoUnit labels on SRD plot
- Conservative assessments: Automatic selection of most conservative values across methods
- Installation risk indicators: Gripper release, MP abandonment, refusal risk depths
- Professional formatting: A3 landscape layout optimized for printing/reporting
- Interactive tooltips: Hover over any point to see exact values
- Robust error handling: Graceful degradation if optional data sources unavailable

DATA SOURCES:
-------------
- Driveability results: results_PileDrivingAnalysis-{Position}.csv (multi-table CSV)
- Soil profiles: input_Position_*_{Position}.csv (qc, layers, geoUnits)
- Monopile weights: Excel file from primary steel design verification
- Gripper penetration: Excel workbook from lateral pile stability analysis
- Refusal risk: Excel workbook with pause duration scenarios (1hr to 7days)

OUTPUT:
-------
- Interactive HTML plots: Installation_Driveability_{Position}.html
- Console output: Progress messages, warnings, loaded data summary
- Plot display: Automatic browser opening for immediate review
"""

import pandas as pd
from pathlib import Path
import plotly.graph_objects as go
import re
import time
from typing import Dict, List, Tuple, Optional
from decimal import Decimal, ROUND_HALF_UP
from plotly.subplots import make_subplots

# ===================================================================================================
# GLOBAL CACHE FOR PARSED FILES (PERFORMANCE OPTIMIZATION)
# ===================================================================================================
_PARSED_FILE_CACHE = {}  # Cache for parsed CSV files to avoid re-parsing


def clear_parse_cache():
    """Clear the parsed file cache to free memory."""
    global _PARSED_FILE_CACHE
    _PARSED_FILE_CACHE.clear()


# ===================================================================================================
# USER CONFIGURABLE CONSTANTS
# ===================================================================================================
# These values control plot thresholds, file locations, and calculation parameters.
# Modify these as needed for your specific project requirements.

# --- DRIVEABILITY THRESHOLD VALUES ---
# Blowcount rate thresholds (units: blows per 25cm)
HARD_DRIVING_BLOWCOUNT = 75   # Hard driving criterion - shown as horizontal line on blowcount plot
REFUSAL_BLOWCOUNT = 250        # Refusal criterion - indicates potential installation failure

# --- WEIGHT COMPONENTS FOR SELF-WEIGHT PENETRATION CALCULATIONS ---
# All weights in tonnes (t)
INTERNAL_LIFTING_TOOL = 100    # Weight of internal lifting tool (ILT) used during MP lowering
HAMMER_WEIGHT = 736            # Weight of hydraulic hammer unit
ADDITIONAL_WEIGHT = 20         # Additional components (flanges, pins, secondary attachments)

# --- DATA SOURCE FILE PATHS ---
# Path to Excel file containing monopile weights for all positions
MONOPILE_WEIGHTS_FILE = Path(r"k:/dozr/HOW03/PS/MP/20250912 - Design documentation for Certification - Rev. C/variations/01_Primary_Steel_Design_Verification_25yr/summary/data_summary/summary-01_Primary_Steel_Design_Verification_25yr.xls")

# Root directory containing position folders (A01, A02, etc.) with driveability results
MONOPILE_ROOT_DIR = Path(r"k:/dozr/HOW03/GEO/05_Driveability/20260106_Final for Installation/variations/02_ConstBlow/monopiles")

# Output directory where interactive HTML plots will be saved
PLOTS_OUTPUT_DIR = Path(r"k:\dozr\HOW03\GEO\05_Driveability\20260106_Final for Installation\variations\02_ConstBlow\summary\post_processing_plots")

# --- GRIPPER PENETRATION DATA SOURCE ---
# Minimum penetration depths required for gripper release and MP abandonment
GRIPPER_PENETRATION_DIR = r"K:\dozr\HOW03\GEO\04_OptiMon Runs\20251017_Lateral_pile_stability_ Installation\post-processing"
GRIPPER_PENETRATION_FILE = "HOW03_minL_load_iter3.xlsm"
GRIPPER_PENETRATION_SHEET = "Summary 05_combined"

# --- REFUSAL RISK ASSESSMENT DATA SOURCE ---
# Refusal risk depths for different installation pause durations (1hr, 24hr, 48hr, 7days)
REFUSAL_RISK_DIR = r"K:\dozr\HOW03\GEO\05_Driveability\20260106_Final for Installation\postprocessing"
REFUSAL_RISK_FILE = "Data_Summary_setup.xlsx"
# ===================================================================================================


def parse_results_csv(file_path: Path) -> Dict[str, pd.DataFrame]:
    """
    Parse a results CSV file containing multiple tables separated by empty rows.

    The CSV files follow a specific format:
    - Each table starts with '**' followed by the table name
    - Line 1 after '**': Table name (e.g., 'results_PileDrivingAnalysis_Summary')
    - Line 2: Position identifier (e.g., 'A01')
    - Line 3: Column headers (semicolon-separated)
    - Line 4: Unit row (ignored)
    - Lines 5+: Data rows (semicolon-separated)

    Handles multiple encoding formats to ensure robust file reading across different systems.

    Args:
        file_path: Path to the CSV file to parse

    Returns:
        Dictionary mapping table names (str) to DataFrames
        Example: {'results_PileDrivingAnalysis_Summary': DataFrame, 'soil': DataFrame, ...}

    Raises:
        ValueError: If file cannot be read with any supported encoding
    """
    # PERFORMANCE OPTIMIZATION: Check cache first
    cache_key = str(file_path)
    if cache_key in _PARSED_FILE_CACHE:
        return _PARSED_FILE_CACHE[cache_key]

    tables = {}

    # Read the entire file - try multiple encodings
    encodings = ['utf-8', 'windows-1252', 'latin-1', 'cp1252', 'iso-8859-1']
    content = None

    for encoding in encodings:
        try:
            with open(str(file_path), 'r', encoding=encoding) as f:
                content = f.read()
            break  # Success - stop trying other encodings
        except (UnicodeDecodeError, LookupError):
            continue  # Try next encoding

    if content is None:
        raise ValueError(f"Could not read file {file_path} with any of the supported encodings: {encodings}")

    # Split by double asterisks to find table headers
    sections = re.split(r'\*\*', content)

    for section in sections[1:]:  # Skip first empty section
        lines = section.strip().split('\n')
        if not lines:
            continue

        # First line is the table name (remove trailing semicolon)
        table_name = lines[0].strip().rstrip(';')

        # Find where the actual data starts (after position line and header)
        # Typically: table_name, position, column_headers, unit_row, data...
        if len(lines) < 4:
            continue

        # Position is on line 1
        position = lines[1].strip()

        # Column headers on line 2
        headers = lines[2].strip().split(';')

        # Skip unit row (line 3) and read data starting from line 4
        data_lines = []
        for line in lines[4:]:
            line = line.strip()
            if not line:  # Stop at empty line (next table)
                break
            data_lines.append(line)

        if not data_lines:
            continue

        # Parse data into DataFrame
        data_rows = [line.split(';') for line in data_lines]
        df = pd.DataFrame(data_rows, columns=headers)

        # OPTIMIZED: Convert numeric columns using vectorized operation
        # Replace common NaN representations once for entire DataFrame
        df = df.replace(['NaN', 'nan', 'NA', 'na', ''], pd.NA)

        # Try to convert columns to numeric (future-proof approach)
        # Iterate through columns but use vectorized conversion per column
        for col in df.columns:
            try:
                df[col] = pd.to_numeric(df[col])
            except (ValueError, TypeError):
                # Keep as string if conversion fails (e.g., plot_Color, geoUnit columns)
                pass

        tables[table_name] = df

    # PERFORMANCE OPTIMIZATION: Store in cache before returning
    _PARSED_FILE_CACHE[cache_key] = tables
    return tables


def get_position_folders(root_dir: Path) -> List[Tuple[str, Path]]:
    """
    Get all position folders (A01, A02, etc.) from the root directory.

    Returns:
        List of tuples (position_name, folder_path)
    """
    positions = []

    if not root_dir.exists():
        print(f"Error: Directory does not exist: {root_dir}")
        return positions

    for item in sorted(root_dir.iterdir()):
        if item.is_dir() and re.match(r'^[A-Z]\d+$', item.name):
            positions.append((item.name, item))

    return positions


def extract_method_and_bound(table_name: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Extract SRD method and soil bound from table name.

    Example: 'results_PileDrivingAnalysis_HOW03_lb_MD_4400S' -> ('MD', 'lb')
    Example: 'results_PileDrivingAnalysis_HOW03_lb_Maynard_4400S' -> ('MY', 'lb')
    """
    # Pattern: ..._{lb|be|ub}_{MD|MY|AH|Maynard}_...
    match = re.search(r'_(lb|be|ub)_(MD|MY|AH|Maynard)_', table_name)
    if match:
        method = match.group(2)
        bound = match.group(1)

        # Map "Maynard" to "MY" for consistency
        if method == "Maynard":
            method = "MY"

        return method, bound
    return None, None


def get_available_methods_and_bounds(tables: Dict[str, pd.DataFrame]) -> Tuple[List[str], List[str]]:
    """
    Extract all unique SRD methods and soil bounds from table names.
    """
    methods = set()
    bounds = set()

    for table_name in tables.keys():
        method, bound = extract_method_and_bound(table_name)
        if method and bound:
            methods.add(method)
            bounds.add(bound)

    return sorted(list(methods)), sorted(list(bounds))


def get_available_methods_and_bounds_from_summary(summary_table: pd.DataFrame) -> Tuple[List[str], List[str], Dict[str, Dict]]:
    """
    Extract all unique mapped SRD methods and soil bounds from the summary table.
    Returns:
        - List of mapped methods (AH, MD, MY)
        - List of soil bounds (lb, be, ub)
        - Dict mapping (method, bound) to analysis_ID and original method
    """
    method_map = {
        "Alm and Hamre": "AH",
        "MonoDrive": "MD",
        "Maynard": "MY"
    }
    methods = set()
    bounds = set()
    method_bound_to_id = {}
    for idx, row in summary_table.iterrows():
        orig_method = str(row.get("method", "")).strip()
        mapped_method = method_map.get(orig_method, None)
        soil_case = str(row.get("SoilCase", "")).strip().lower()
        analysis_id = str(row.get("analysis_ID", "")).strip()
        if mapped_method and soil_case:
            methods.add(mapped_method)
            bounds.add(soil_case)
            method_bound_to_id[(mapped_method, soil_case)] = {
                "analysis_ID": analysis_id,
                "orig_method": orig_method
            }
    return sorted(list(methods)), sorted(list(bounds)), method_bound_to_id


def get_monopile_weights(excel_path: Path, positions: list) -> dict:
    """
    Reads the Excel file and returns a dict mapping position name to monopile weight for the given positions.
    Combines rows 7 and 8 for headers, starts data from row 10.
    Uses exact column names 'Position' and 'MP mass'.

    IMPROVED VERSION with better error handling and more flexible column matching.
    """
    try:
        df_raw = pd.read_excel(excel_path, sheet_name=0, header=None)

        # Combine header rows (rows 7 and 8, 0-indexed as 6 and 7)
        header_rows = df_raw.iloc[6:8].astype(str)
        combined_headers = header_rows.agg(' '.join).str.replace('nan ', '', regex=False).str.strip()
        df_raw.columns = combined_headers

        # Data starts at row 10 (0-indexed as 9)
        df = df_raw.iloc[9:].reset_index(drop=True)

        # Normalize column names for flexible matching
        col_mapping = {col: col.strip().lower() for col in df.columns}

        # Find position and weight columns (case-insensitive and flexible matching)
        position_col = next((col for col, norm in col_mapping.items() if norm == 'position'), None)
        weight_col = next((col for col, norm in col_mapping.items() if 'mp mass' in norm or 'mp_mass' in norm), None)

        if position_col is None or weight_col is None:
            print(f"ERROR: Could not find required columns for position or MP mass.")
            print(f"Available columns: {list(df.columns)}")
            return {}

        # Build weights dictionary
        weights = {}
        for _, row in df.iterrows():
            pos = str(row[position_col]).strip()
            if pos in positions:
                try:
                    weight = float(row[weight_col])
                except (ValueError, TypeError):
                    weight = None
                weights[pos] = weight

        return weights

    except Exception as e:
        print(f"ERROR reading monopile weights from {excel_path}: {e}")
        return {}


def get_position_info(position_tables, position, selected_methods=None, selected_bounds=None):
    """
    Extracts hammer name, hammer weight, target blowcount rate, and target penetration depth for a given position.
    Returns a dict with keys: 'hammer_name', 'hammer_weight', 'target_blowcount_rate', 'target_penetration_depth'.

    If selected_methods and selected_bounds are provided, checks that all selected
    method/bound combinations use the same hammer. If different hammers are used,
    returns 'different hammers used'.
    """
    # Hammer weight always from user constant
    info = {'hammer_name': None, 'hammer_weight': HAMMER_WEIGHT, 'target_blowcount_rate': None, 'target_penetration_depth': None}
    # position_tables is already the tables dict for this position
    tables = position_tables

    # Get hammer name from summary table
    summary_table = tables.get('results_PileDrivingAnalysis_Summary')
    if summary_table is not None and 'Hammer_name' in summary_table.columns:
        # If selected methods and bounds are provided, check for consistency
        if selected_methods is not None and selected_bounds is not None:
            hammer_names = set()

            # Check hammer name for each selected method/bound combination
            for idx, row in summary_table.iterrows():
                # Extract method and bound from this row
                method_col = row.get('Method', '')
                bound_col = row.get('SoilCase', '')

                # Map full method names to short codes
                method_short = None
                if 'MonoDrive' in str(method_col):
                    method_short = 'MD'
                elif 'Maynard' in str(method_col):
                    method_short = 'MY'
                elif 'Alm and Hamre' in str(method_col) or 'Alm' in str(method_col):
                    method_short = 'AH'

                # Check if this row matches selected methods and bounds
                if method_short in selected_methods and str(bound_col).lower() in selected_bounds:
                    hammer_name = row.get('Hammer_name', None)
                    if pd.notna(hammer_name):
                        hammer_names.add(str(hammer_name))

            # Check if all selected combinations have the same hammer
            if len(hammer_names) == 0:
                info['hammer_name'] = ''
            elif len(hammer_names) == 1:
                info['hammer_name'] = hammer_names.pop()
            else:
                info['hammer_name'] = 'different hammers used'
        else:
            # No filtering - just get first hammer name
            val = summary_table['Hammer_name'].iloc[0]
            if pd.notna(val):
                info['hammer_name'] = str(val)

    # Target blowcount rate from summary table column 'Target_Blowcount_Rate'
    # Note: CSV values are in blows/m, but we display in blows/25cm, so divide by 4
    if summary_table is not None and 'Target_Blowcount_Rate' in summary_table.columns:
        val = summary_table['Target_Blowcount_Rate'].iloc[0]
        if pd.notna(val):
            info['target_blowcount_rate'] = float(val) / 4.0  # Convert blows/m to blows/25cm

    # Target penetration depth from summary table column 'targetdepth'
    if summary_table is not None and 'targetdepth' in summary_table.columns:
        val = summary_table['targetdepth'].iloc[0]
        if pd.notna(val):
            info['target_penetration_depth'] = float(val)

    return info


def get_gripper_penetration_for_positions(positions: List[str]) -> tuple[Dict[str, float], Dict[str, float]]:
    """
    Read minimum penetration data from Excel file for SPECIFIC positions only.

    Reads two types of penetration data:
    1. Gripper release: Lmin_Hs2_5 (SLS,ULS)
    2. MP abandonment: Lmin_Hs7_3 (SLS,ULS)

    This function is called AFTER position selection to read only the required data,
    making it more efficient than loading all positions upfront.

    Args:
        positions: List of position names to read data for (e.g., ['A01', 'A02'])

    Returns:
        Tuple of two dictionaries:
        - gripper_data: Dictionary mapping position names to gripper release penetration depths
        - abandonment_data: Dictionary mapping position names to MP abandonment penetration depths
        Returns empty dictionaries if file not found or any error occurs.
    """
    if not positions:
        return {}, {}

    try:
        # Construct full path to Excel file
        excel_path = Path(GRIPPER_PENETRATION_DIR) / GRIPPER_PENETRATION_FILE

        if not excel_path.exists():
            print(f"Info: Penetration data file not found at {excel_path}. Skipping penetration data.")
            return {}, {}

        print(f"Reading penetration data for {len(positions)} selected position(s)...")

        # Read the Excel file, starting from row 18 (0-indexed row 17)
        # Row 18 contains headers, row 19 contains units, data starts at row 20
        df = pd.read_excel(excel_path, sheet_name=GRIPPER_PENETRATION_SHEET, header=17)

        # Remove the units row (first row after header)
        df = df.iloc[1:].reset_index(drop=True)

        # Check if required columns exist
        if 'Position' not in df.columns:
            print(f"Warning: 'Position' column not found in {GRIPPER_PENETRATION_SHEET}. Skipping penetration data.")
            return {}, {}

        # Column names may have newline character due to Excel formatting
        # Find gripper release column (Hs2_5)
        gripper_col = 'Lmin_Hs2_5 (SLS,ULS)'
        if gripper_col not in df.columns:
            # Try with newline character
            gripper_col = 'Lmin_Hs2_5\n(SLS,ULS)'
            if gripper_col not in df.columns:
                print(f"Warning: Column 'Lmin_Hs2_5 (SLS,ULS)' not found in {GRIPPER_PENETRATION_SHEET}.")
                gripper_col = None

        # Find MP abandonment column (Hs7_3)
        abandonment_col = 'Lmin_Hs7_3 (SLS,ULS)'
        if abandonment_col not in df.columns:
            # Try with newline character
            abandonment_col = 'Lmin_Hs7_3\n(SLS,ULS)'
            if abandonment_col not in df.columns:
                print(f"Warning: Column 'Lmin_Hs7_3 (SLS,ULS)' not found in {GRIPPER_PENETRATION_SHEET}.")
                abandonment_col = None

        if not gripper_col and not abandonment_col:
            print(f"Warning: No penetration columns found. Skipping penetration data.")
            return {}, {}

        # Filter dataframe to only include selected positions
        df_filtered = df[df['Position'].isin(positions)]

        # Build dictionaries of position -> penetration depth (only for selected positions)
        gripper_data = {}
        abandonment_data = {}

        for idx, row in df_filtered.iterrows():
            pos = row.get('Position')

            if pd.notna(pos):
                # Read gripper release data
                if gripper_col:
                    val = row.get(gripper_col)
                    if pd.notna(val):
                        try:
                            gripper_data[str(pos)] = float(val)
                        except (ValueError, TypeError):
                            pass  # Skip if value cannot be converted

                # Read MP abandonment data
                if abandonment_col:
                    val = row.get(abandonment_col)
                    if pd.notna(val):
                        try:
                            abandonment_data[str(pos)] = float(val)
                        except (ValueError, TypeError):
                            pass  # Skip if value cannot be converted

        # Report what was loaded
        if gripper_data or abandonment_data:
            msg_parts = []
            if gripper_data:
                msg_parts.append(f"gripper release: {len(gripper_data)}/{len(positions)}")
            if abandonment_data:
                msg_parts.append(f"MP abandonment: {len(abandonment_data)}/{len(positions)}")
            print(f"  ✓ Loaded penetration data - {', '.join(msg_parts)}")
        else:
            print(f"  Warning: No valid penetration data found for selected positions")

        return gripper_data, abandonment_data

    except Exception as e:
        print(f"Warning: Error reading penetration data: {e}. Skipping penetration data.")
        return {}, {}


def get_refusal_risk_for_positions(positions: List[str], selected_methods: List[str],
                                    target_depths: Dict[str, float]) -> Dict[str, Dict[str, float]]:
    """
    Read refusal risk depth data from Excel file for specific positions and installation pause durations.

    Reads refusal risk assessment data from sheets named: '1 hr', '24 hr', '48 hr', '7 days'

    For each sheet:
    - First 3 rows are header (row 1: Position/difference, row 2: all/<SRD method>, row 3: text/<soil bound>)
    - Data starts from row 4 onwards
    - Find the position row and the column matching the selected SRD method
    - The value represents how far ABOVE target depth the refusal occurs (negative number)
    - Calculate actual refusal depth as: target_depth + difference

    Args:
        positions: List of position names to read data for
        selected_methods: List of selected SRD methods (e.g., ['MD', 'AH'])
        target_depths: Dictionary mapping position names to target penetration depths

    Returns:
        Nested dictionary:
        {
            'position_name': {
                '1hr': depth_value,
                '24hr': depth_value,
                '48hr': depth_value,
                '7days': depth_value
            }
        }
        Returns empty dict if file not found or any error occurs.
    """
    if not positions:
        return {}

    try:
        # Construct full path to Excel file
        excel_path = Path(REFUSAL_RISK_DIR) / REFUSAL_RISK_FILE

        if not excel_path.exists():
            print(f"Info: Refusal risk data file not found at {excel_path}. Skipping refusal risk data.")
            return {}

        print(f"Reading refusal risk data for {len(positions)} selected position(s)...")

        # Sheet names mapping
        sheet_mapping = {
            '1hr': '1 hr',
            '24hr': '24 hr',
            '48hr': '48 hr',
            '7days': '7 days'
        }

        # Method name mapping (from short codes to expected column headers)
        method_mapping = {
            'MD': 'MD',
            'MY': 'Maynard',
            'AH': 'AH'
        }

        # Initialize results dictionary
        refusal_data = {pos: {} for pos in positions}

        # Read each sheet
        for duration_key, sheet_name in sheet_mapping.items():
            try:
                # Read the sheet (first 3 rows are header)
                df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)

                if df.empty or len(df) < 4:
                    print(f"  Warning: Sheet '{sheet_name}' is empty or too short. Skipping.")
                    continue

                # Row 0: Contains 'Position' in first column, 'difference' and other headers
                # Row 1: Contains 'all' in first column, SRD method names in other columns
                # Row 2: Contains 'text' in first column, soil bounds (be/lb/ub) in other columns
                # Row 3+: Data rows

                header_row1 = df.iloc[0]  # Position / difference
                header_row2 = df.iloc[1]  # all / SRD methods
                header_row3 = df.iloc[2]  # text / soil bounds

                # Find 'Position' column (should be first column)
                position_col_idx = 0

                # Find columns that match selected methods (any soil bound)
                # We look in row 1 for method names and check if they match our selected methods
                method_columns = {}  # {method_short: [col_indices]}

                for col_idx in range(1, len(header_row2)):
                    method_in_header = str(header_row2.iloc[col_idx]).strip()

                    # Check if this column matches any of our selected methods
                    for method_short, method_name in method_mapping.items():
                        if method_short in selected_methods:
                            if method_name.lower() in method_in_header.lower():
                                if method_short not in method_columns:
                                    method_columns[method_short] = []
                                method_columns[method_short].append(col_idx)

                if not method_columns:
                    print(f"  Warning: No matching methods found in sheet '{sheet_name}'. Skipping.")
                    continue

                # Process each position
                for pos in positions:
                    # Find the row for this position (starting from row 3, index 3)
                    position_row_idx = None
                    for row_idx in range(3, len(df)):
                        if str(df.iloc[row_idx, position_col_idx]).strip() == pos:
                            position_row_idx = row_idx
                            break

                    if position_row_idx is None:
                        continue  # Position not found in this sheet

                    # Get target depth for this position
                    target_depth = target_depths.get(pos)
                    if target_depth is None:
                        continue  # No target depth available

                    # Extract difference values for all selected methods
                    # Take the first matching column for each method (typically 'be' bound)
                    differences = []
                    for method_short, col_indices in method_columns.items():
                        if col_indices:
                            val = df.iloc[position_row_idx, col_indices[0]]
                            if pd.notna(val):
                                try:
                                    diff_value = float(val)
                                    differences.append(diff_value)
                                except (ValueError, TypeError):
                                    pass

                    # If we have valid differences, take the MOST CONSERVATIVE (most negative/shallowest)
                    # This represents the earliest refusal risk across all selected SRD methods
                    # The difference is negative (distance above target), so we add it to target
                    # Only show "No risk" if ALL methods show 0 difference (all safe)
                    if differences:
                        # Take the minimum (most negative) difference = most conservative = earliest refusal
                        most_conservative_diff = min(differences)
                        # Only mark as "No risk" if the most conservative value is 0
                        # (meaning all values are >= 0, i.e., all methods show no risk)
                        if most_conservative_diff == 0:
                            refusal_data[pos][duration_key] = "No risk"
                        else:
                            # Calculate refusal depth using most conservative difference
                            refusal_depth = target_depth + most_conservative_diff
                            refusal_data[pos][duration_key] = refusal_depth

            except Exception as e:
                print(f"  Warning: Error reading sheet '{sheet_name}': {e}")
                continue

        # Report what was loaded
        loaded_count = sum(1 for pos_data in refusal_data.values() if pos_data)
        if loaded_count > 0:
            print(f"  ✓ Loaded refusal risk data for {loaded_count}/{len(positions)} position(s)")
        else:
            print(f"  Warning: No valid refusal risk data found for selected positions")

        return refusal_data

    except Exception as e:
        print(f"Warning: Error reading refusal risk data: {e}. Skipping refusal risk data.")
        return {}


def calculate_swp_and_pile_run_assessment(
    methods_to_plot: List[str],
    bounds_to_plot: List[str],
    ruts_to_plot: List,
    depths_to_plot: List,
    mp_weight: Optional[float]
) -> dict:
    """
    Calculate self-weight penetration (SWP) and pile run assessment metrics.

    This function evaluates installation risks by determining depths where the monopile
    may penetrate under its own weight or experience pile run (uncontrolled descent).

    CONSERVATIVE VALUE SELECTION LOGIC (when multiple SRD methods are selected):
    - SWP MP + ILT: DEEPEST depth → Pile penetrates furthest (most conservative)
    - Pile run at hammer placement: SHALLOWEST depth → Earliest risk occurrence
    - SWP MP + Hammer: DEEPEST depth → Pile penetrates furthest with hammer
    - Pile run risk top: SHALLOWEST depth → Earliest risk initiation
    - Pile run risk bottom: DEEPEST depth → Longest risk zone extent

    Weight Components:
    - Nominal MP weight = MP weight + ADDITIONAL_WEIGHT (flanges, pins, etc.)
    - MP + ILT = Nominal MP weight + INTERNAL_LIFTING_TOOL
    - MP + Hammer = Nominal MP weight + HAMMER_WEIGHT

    Assessment Metrics:
    1. SWP MP + ILT: Depth where SRD equals MP+ILT weight (lowering phase)
    2. Pile run at hammer placement: Check if SRD < threshold at SWP MP+ILT depth
    3. SWP MP + Hammer: Depth where SRD equals MP+Hammer weight (driving phase)
    4. Pile run risk top: First depth where SRD < MP+Hammer weight (risk initiation)
    5. Pile run risk bottom: Depth where SRD recovers > MP+Hammer weight (risk zone end)

    Args:
        methods_to_plot: List of SRD method names for each plotted line (e.g., ['MD', 'AH', 'MD'])
        bounds_to_plot: List of bound names for each plotted line (e.g., ['lb', 'be', 'ub'])
        ruts_to_plot: List of SRD arrays in MN for each plotted line
        depths_to_plot: List of depth arrays in meters for each plotted line
        mp_weight: Monopile weight in tonnes (None if unavailable)

    Returns:
        Dictionary containing:
            - 'swp_mp_ilt_depths': Dict with LB, BE, UB keys → depths or 'No penetration'
            - 'pile_run_at_hammer_placement': Dict with LB, BE, UB → 'Yes', 'No', or ''
            - 'swp_mp_hammer_depths': Dict with LB, BE, UB → depths or 'No penetration'
            - 'pile_run_risk_top': Dict with LB, BE, UB → depths or 'No risk'
            - 'pile_run_risk_bottom': Dict with LB, BE, UB → depths, 'No bottom', or 'No risk'
            - 'nominal_mp_weight': Nominal MP weight in tonnes
            - 'mp_lift_tool_total_weight_kn': MP+ILT weight in kN
            - 'mp_hammer_total_weight_kn': MP+Hammer weight in kN
            - 'mp_only_total_weight_kn': Nominal MP weight in kN
    """
    # Initialize result dictionary
    results = {
        'swp_mp_ilt_depths': {'LB': '', 'BE': '', 'UB': ''},
        'pile_run_at_hammer_placement': {'LB': '', 'BE': '', 'UB': ''},
        'swp_mp_hammer_depths': {'LB': '', 'BE': '', 'UB': ''},
        'pile_run_risk_top': {'LB': '', 'BE': '', 'UB': ''},
        'pile_run_risk_bottom': {'LB': '', 'BE': '', 'UB': ''},
        'nominal_mp_weight': None,
        'mp_lift_tool_total_weight_kn': None,
        'mp_hammer_total_weight_kn': None,
        'mp_only_total_weight_kn': None
    }

    # Return empty results if no monopile weight provided
    if mp_weight is None:
        return results

    # Calculate nominal monopile weight (MP + additional weight)
    nominal_mp_weight = mp_weight + ADDITIONAL_WEIGHT
    results['nominal_mp_weight'] = nominal_mp_weight

    # Calculate weight thresholds in kN
    mp_lift_tool_total_weight_kn = (nominal_mp_weight + INTERNAL_LIFTING_TOOL) * 9.81
    mp_hammer_total_weight_kn = (nominal_mp_weight + HAMMER_WEIGHT) * 9.81
    mp_only_total_weight_kn = nominal_mp_weight * 9.81

    results['mp_lift_tool_total_weight_kn'] = mp_lift_tool_total_weight_kn
    results['mp_hammer_total_weight_kn'] = mp_hammer_total_weight_kn
    results['mp_only_total_weight_kn'] = mp_only_total_weight_kn

    # Store numeric values for later calculations
    swp_mp_ilt_depths_numeric = {'LB': None, 'BE': None, 'UB': None}
    swp_mp_hammer_depths_numeric = {'LB': None, 'BE': None, 'UB': None}

    # CONSERVATIVE VALUE SELECTION LOGIC:
    # When multiple SRD methods are selected, we take the most conservative value for each metric:
    # - SWP MP + ILT: DEEPEST depth (pile penetrates furthest under its own weight)
    # - Pile run at hammer placement: SHALLOWEST depth (earliest/highest risk occurrence)
    # - SWP MP + Hammer: DEEPEST depth (pile penetrates furthest with hammer weight)
    # - Pile run risk top: SHALLOWEST depth (earliest risk initiation)
    # - Pile run risk bottom: DEEPEST depth (longest risk zone extent)

    # Temporary storage for all calculated values across all methods
    temp_swp_ilt = {'LB': [], 'BE': [], 'UB': []}
    temp_pile_run_hammer_placement = {'LB': [], 'BE': [], 'UB': []}
    temp_swp_hammer = {'LB': [], 'BE': [], 'UB': []}
    temp_pile_run_top = {'LB': [], 'BE': [], 'UB': []}
    temp_pile_run_bottom = {'LB': [], 'BE': [], 'UB': []}

    # --- Calculate SWP MP + ILT depths for each bound ---
    # LOGIC: Take DEEPEST depth across all methods (most conservative for SWP)
    for method, bound, rut, depth in zip(methods_to_plot, bounds_to_plot, ruts_to_plot, depths_to_plot):
        rut_kn = rut * 1000
        bound_key = bound.upper()

        # Find first crossing with MP+ILT weight
        for i in range(1, len(rut_kn)):
            if (rut_kn[i-1] < mp_lift_tool_total_weight_kn <= rut_kn[i]) or \
               (rut_kn[i-1] > mp_lift_tool_total_weight_kn >= rut_kn[i]):
                # Linear interpolation for depth
                d1, d2 = depth[i-1], depth[i]
                r1, r2 = rut_kn[i-1], rut_kn[i]
                if r2 != r1:
                    depth_cross = d1 + (mp_lift_tool_total_weight_kn - r1) * (d2 - d1) / (r2 - r1)
                else:
                    depth_cross = d1
                temp_swp_ilt[bound_key].append(depth_cross)
                break

    # Select DEEPEST depth for each bound (most conservative for SWP)
    for bound_key in ['LB', 'BE', 'UB']:
        if temp_swp_ilt[bound_key]:
            deepest = max(temp_swp_ilt[bound_key])
            results['swp_mp_ilt_depths'][bound_key] = f'{deepest:.2f}'
            swp_mp_ilt_depths_numeric[bound_key] = deepest

    # --- Calculate pile run at hammer placement (SRD < MP only) ---
    # LOGIC: Take SHALLOWEST depth across all methods (earliest/highest risk - most conservative)
    for method, bound, rut, depth in zip(methods_to_plot, bounds_to_plot, ruts_to_plot, depths_to_plot):
        rut_kn = rut * 1000
        bound_key = bound.upper()

        for i in range(1, len(rut_kn)):
            if rut_kn[i-1] >= mp_only_total_weight_kn and rut_kn[i] < mp_only_total_weight_kn:
                d1, d2 = depth[i-1], depth[i]
                r1, r2 = rut_kn[i-1], rut_kn[i]
                if r2 != r1:
                    depth_cross = d1 + (mp_only_total_weight_kn - r1) * (d2 - d1) / (r2 - r1)
                else:
                    depth_cross = d1
                temp_pile_run_hammer_placement[bound_key].append(depth_cross)
                break

    # Select SHALLOWEST depth for each bound (earliest risk - most conservative)
    for bound_key in ['LB', 'BE', 'UB']:
        if temp_pile_run_hammer_placement[bound_key]:
            shallowest = min(temp_pile_run_hammer_placement[bound_key])
            results['pile_run_at_hammer_placement'][bound_key] = f'{shallowest:.2f}'
        else:
            results['pile_run_at_hammer_placement'][bound_key] = 'No risk'

    # --- Calculate SWP MP + Hammer depths ---
    # LOGIC: Take DEEPEST depth across all methods (most conservative for SWP)
    for method, bound, rut, depth in zip(methods_to_plot, bounds_to_plot, ruts_to_plot, depths_to_plot):
        rut_kn = rut * 1000
        bound_key = bound.upper()

        # Get numeric values of first two rows to determine start depth
        depth1 = swp_mp_ilt_depths_numeric.get(bound_key, None)
        depth2_str = results['pile_run_at_hammer_placement'].get(bound_key, None)
        try:
            depth2_num = float(depth2_str) if depth2_str and depth2_str != 'No risk' else None
        except Exception:
            depth2_num = None

        # Compute start depth (max of the two)
        start_depth = None
        if depth1 is not None and depth2_num is not None:
            start_depth = max(depth1, depth2_num)
        elif depth1 is not None:
            start_depth = depth1
        elif depth2_num is not None:
            start_depth = depth2_num

        # Find first upward crossing after start_depth
        if start_depth is not None:
            for i in range(1, len(rut_kn)):
                if depth[i] > start_depth:
                    # Only upward crossing: SRD rises above MP+Hammer total weight
                    if rut_kn[i-1] < mp_hammer_total_weight_kn <= rut_kn[i]:
                        d1, d2 = depth[i-1], depth[i]
                        r1, r2 = rut_kn[i-1], rut_kn[i]
                        if r2 != r1:
                            depth_cross = d1 + (mp_hammer_total_weight_kn - r1) * (d2 - d1) / (r2 - r1)
                        else:
                            depth_cross = d1
                        temp_swp_hammer[bound_key].append(depth_cross)
                        break

    # Select DEEPEST depth for each bound (most conservative for SWP)
    for bound_key in ['LB', 'BE', 'UB']:
        if temp_swp_hammer[bound_key]:
            deepest = max(temp_swp_hammer[bound_key])
            results['swp_mp_hammer_depths'][bound_key] = f'{deepest:.2f}'
            swp_mp_hammer_depths_numeric[bound_key] = deepest

    # --- Calculate pile run risk top (initiation) ---
    # LOGIC: Take SHALLOWEST depth across all methods (earliest risk initiation - most conservative)
    for method, bound, rut, depth in zip(methods_to_plot, bounds_to_plot, ruts_to_plot, depths_to_plot):
        rut_kn = rut * 1000
        bound_key = bound.upper()
        swp_mp_hammer_depth = swp_mp_hammer_depths_numeric.get(bound_key, None)

        if swp_mp_hammer_depth is not None:
            for i in range(1, len(rut_kn)):
                current_depth = depth[i]
                # Only consider depths below the SWP MP+Hammer depth
                if current_depth > swp_mp_hammer_depth:
                    if rut_kn[i] < mp_hammer_total_weight_kn:
                        # Linear interpolation to find exact crossing depth
                        if i > 0 and rut_kn[i-1] >= mp_hammer_total_weight_kn:
                            d1, d2 = depth[i-1], depth[i]
                            r1, r2 = rut_kn[i-1], rut_kn[i]
                            if r2 != r1:
                                depth_cross = d1 + (mp_hammer_total_weight_kn - r1) * (d2 - d1) / (r2 - r1)
                            else:
                                depth_cross = current_depth
                        else:
                            depth_cross = current_depth
                        temp_pile_run_top[bound_key].append(depth_cross)
                        break

    # Select SHALLOWEST depth for each bound (earliest risk initiation - most conservative)
    for bound_key in ['LB', 'BE', 'UB']:
        if temp_pile_run_top[bound_key]:
            shallowest = min(temp_pile_run_top[bound_key])
            results['pile_run_risk_top'][bound_key] = f'{shallowest:.2f}'
        else:
            results['pile_run_risk_top'][bound_key] = 'No risk'

    # --- Calculate pile run risk bottom ---
    # LOGIC: Take DEEPEST depth across all methods (longest risk zone - most conservative)
    for method, bound, rut, depth in zip(methods_to_plot, bounds_to_plot, ruts_to_plot, depths_to_plot):
        rut_kn = rut * 1000
        bound_key = bound.upper()

        # Get numeric value for pile run risk top
        risk_top_str = results['pile_run_risk_top'].get(bound_key, None)
        if risk_top_str == 'No risk':
            continue  # Skip this method for this bound

        try:
            risk_top = float(risk_top_str) if risk_top_str else None
        except Exception:
            risk_top = None

        if risk_top is not None:
            # Find first index deeper than risk_top
            for i in range(len(depth)):
                if depth[i] > risk_top:
                    # Check if all remaining SRD values are above threshold
                    all_above = all(r > mp_hammer_total_weight_kn for r in rut_kn[i:])
                    if all_above:
                        temp_pile_run_bottom[bound_key].append(depth[i])
                        break

    # Select DEEPEST depth for each bound (longest risk zone - most conservative)
    for bound_key in ['LB', 'BE', 'UB']:
        if temp_pile_run_bottom[bound_key]:
            deepest = max(temp_pile_run_bottom[bound_key])
            results['pile_run_risk_bottom'][bound_key] = f'{deepest:.2f}'
        elif results['pile_run_risk_top'][bound_key] == 'No risk':
            results['pile_run_risk_bottom'][bound_key] = 'No risk'
        else:
            results['pile_run_risk_bottom'][bound_key] = 'No bottom'

    return results


def _parse_soil_tables_only(file_path: Path):
    """
    FAST: Extract only **soil and **position_data tables from input CSV file.

    This is a performance-optimized version that only parses the two tables needed
    for soil profile visualization, instead of parsing the entire file.

    Returns:
        dict with 'soil' and 'position_data' DataFrames, or None on error
    """
    try:
        # Read file with first successful encoding
        encodings = ['utf-8', 'windows-1252', 'latin-1']
        content = None
        for encoding in encodings:
            try:
                with open(str(file_path), 'r', encoding=encoding) as f:
                    content = f.read()
                break
            except (UnicodeDecodeError, LookupError):
                continue

        if content is None:
            return None

        # Only extract the tables we need
        tables = {}
        sections = re.split(r'\*\*', content)

        for section in sections[1:]:
            lines = section.strip().split('\n')
            if not lines:
                continue

            table_name = lines[0].strip().rstrip(';')

            # OPTIMIZATION: Only process soil and position_data tables
            if table_name not in ('soil', 'position_data'):
                continue

            if len(lines) < 4:
                continue

            headers = lines[2].strip().split(';')
            data_lines = []
            for line in lines[4:]:
                line = line.strip()
                if not line:
                    break
                data_lines.append(line)

            if not data_lines:
                continue

            # Parse data
            data_rows = [line.split(';') for line in data_lines]
            df = pd.DataFrame(data_rows, columns=headers)

            # OPTIMIZATION: Only convert numeric columns we actually need
            if table_name == 'soil':
                # Only convert z_top and qc to numeric (skip plot_Color, geoUnit)
                for col in ['z_top', 'qc']:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col].replace(['NaN', 'nan', 'NA', 'na', ''], pd.NA), errors='coerce')
            elif table_name == 'position_data':
                # Only convert zSeabed
                if 'zSeabed' in df.columns:
                    df['zSeabed'] = pd.to_numeric(df['zSeabed'].replace(['NaN', 'nan', 'NA', 'na', ''], pd.NA), errors='coerce')

            tables[table_name] = df

            # Early exit if we have both tables
            if 'soil' in tables and 'position_data' in tables:
                break

        return tables if tables else None
    except Exception as e:
        return None


def extract_soil_profile_data(input_file_path: Path):
    """
    Extract soil profile data from input position CSV file for overlay on driveability plots.

    Reads the **soil and **position_data tables to create a complete soil stratigraphy profile
    with cone resistance (qc) values, layer colors, and geoUnit classifications.

    Data Processing:
    1. Read zSeabed elevation from position_data table (LAT reference)
    2. Extract soil layers from soil table (z_top, qc, plot_Color, geoUnit)
    3. Convert elevations to depths below seabed: depth = zSeabed - z_top
    4. Convert qc from Pa to MPa for plotting
    5. Create step plot arrays for constant qc within each layer
    6. Group consecutive layers with same geoUnit for zone labels

    Filtering:
    - Only includes layers at or below seabed (depth >= 0)
    - Handles missing colors with default gray (#d3d3d3)
    - Processes geoUnit labels if column exists

    Args:
        input_file_path: Path to input_Position_*_{Position}.csv file

    Returns:
        Dictionary containing:
        - 'depth_below_seabed': pd.Series of depths in meters (0 = seabed)
        - 'qc_mpa': pd.Series of cone resistance values in MPa
        - 'colors': pd.Series of hex color codes for each layer
        - 'geo_units': pd.Series of geoUnit names (or None if column missing)
        - 'qc_step': List of qc values for step plot (creates horizontal lines)
        - 'depth_step': List of depth values for step plot (matches qc_step)
        - 'max_qc': Maximum qc value across all layers (for axis scaling)
        - 'geo_unit_zones': List of tuples (unit_name, start_depth, end_depth)

        Returns None if data cannot be extracted or file cannot be parsed.

    Example:
        data = extract_soil_profile_data(Path("input_Position_const_blow_A01.csv"))
        if data:
            print(f"Extracted {len(data['depth_below_seabed'])} soil layers")
            print(f"Max qc: {data['max_qc']:.2f} MPa")
    """
    # PERFORMANCE OPTIMIZATION: Use fast targeted parser instead of full file parser
    try:
        tables = _parse_soil_tables_only(input_file_path)
        if tables is None:
            return None
    except Exception as e:
        print(f"Error parsing input file {input_file_path}: {e}")
        return None

    # Get zSeabed from position_data table
    position_data_table = tables.get('position_data')
    zSeabed = None
    if position_data_table is not None and 'zSeabed' in position_data_table.columns:
        try:
            zSeabed_value = position_data_table['zSeabed'].iloc[0]
            zSeabed = pd.to_numeric(zSeabed_value, errors='coerce')
            if pd.isna(zSeabed):
                zSeabed = None
        except (IndexError, KeyError):
            zSeabed = None

    # Get soil table
    soil_table = tables.get('soil')
    if soil_table is None:
        return None

    # Extract required columns - check existence first
    required_cols = ['z_top', 'qc', 'plot_Color']
    if not all(col in soil_table.columns for col in required_cols):
        return None

    # OPTIMIZED: Extract columns directly (parse_results_csv already converted to numeric)
    z_top_lat = soil_table['z_top']
    qc_pa = soil_table['qc']

    # Handle colors efficiently
    colors = soil_table['plot_Color'].copy()
    if colors.dtype == 'object':  # Only process if not already numeric
        colors = colors.fillna('#d3d3d3').astype(str)
        colors = colors.replace(['nan', 'NaN', 'NA', ''], '#d3d3d3')

    # Extract geoUnit if available
    geo_units = None
    if 'geoUnit' in soil_table.columns:
        geo_units = soil_table['geoUnit'].fillna('').astype(str)

    # OPTIMIZED: Single boolean mask for all filtering
    valid_initial = z_top_lat.notna() & qc_pa.notna()

    # Apply mask once to all arrays
    z_top_lat = z_top_lat[valid_initial].reset_index(drop=True)
    qc_pa = qc_pa[valid_initial].reset_index(drop=True)
    colors = colors[valid_initial].reset_index(drop=True)
    if geo_units is not None:
        geo_units = geo_units[valid_initial].reset_index(drop=True)

    if len(z_top_lat) == 0:
        return None

    # Determine seabed elevation
    seabed_elevation = zSeabed if zSeabed is not None else z_top_lat.iloc[0]

    # OPTIMIZED: Vectorized calculations
    depth_below_seabed = seabed_elevation - z_top_lat
    qc_mpa = qc_pa / 1e6

    # OPTIMIZED: Single combined filter
    valid_mask = (depth_below_seabed >= 0) & (qc_mpa > 0)
    depth_below_seabed = depth_below_seabed[valid_mask].reset_index(drop=True)
    qc_mpa = qc_mpa[valid_mask].reset_index(drop=True)
    colors = colors[valid_mask].reset_index(drop=True)
    if geo_units is not None:
        geo_units = geo_units[valid_mask].reset_index(drop=True)

    if len(depth_below_seabed) == 0:
        return None

    # OPTIMIZED: Build step plot data using list comprehension (faster than extend in loop)
    n = len(depth_below_seabed)
    if n > 1:
        # Vectorized approach for step plot
        qc_step = []
        depth_step = []
        for i in range(n - 1):
            qc_val = qc_mpa.iloc[i]
            d_start = depth_below_seabed.iloc[i]
            d_end = depth_below_seabed.iloc[i + 1]
            qc_step.extend([qc_val, qc_val])
            depth_step.extend([d_start, d_end])
        # Add last point
        qc_step.append(qc_mpa.iloc[-1])
        depth_step.append(depth_below_seabed.iloc[-1])
    else:
        qc_step = [qc_mpa.iloc[0]]
        depth_step = [depth_below_seabed.iloc[0]]

    max_qc = float(qc_mpa.max())  # Convert to Python float (faster than keeping as numpy)

    # OPTIMIZED: Identify geoUnit zones efficiently
    geo_unit_zones = []
    if geo_units is not None and len(geo_units) > 0:
        current_unit = None
        zone_start = None

        for i in range(len(geo_units)):
            unit = geo_units.iloc[i]
            depth = depth_below_seabed.iloc[i]

            # Skip empty/invalid units
            if not unit or unit in ('', 'nan', 'NaN', 'NA'):
                if current_unit is not None:
                    geo_unit_zones.append((current_unit, zone_start, depth))
                    current_unit = None
                    zone_start = None
                continue

            # Start new zone or continue existing
            if unit != current_unit:
                if current_unit is not None:
                    geo_unit_zones.append((current_unit, zone_start, depth))
                current_unit = unit
                zone_start = depth

        # Close last zone
        if current_unit is not None:
            last_depth = depth_below_seabed.iloc[-1]
            geo_unit_zones.append((current_unit, zone_start, last_depth))

    return {
        'depth_below_seabed': depth_below_seabed,
        'qc_mpa': qc_mpa,
        'colors': colors,
        'geo_units': geo_units,
        'qc_step': qc_step,
        'depth_step': depth_step,
        'max_qc': max_qc,
        'geo_unit_zones': geo_unit_zones
    }


def plot_soil_profile(position: str, input_file_path: Path, output_dir: Path = None):
    """
    Plot soil profile: Depth below mudline vs qc (cone resistance).

    This function is COMPLETELY INDEPENDENT from the driveability plotting.
    It reads its own input file, processes its own data, and creates its own plot.

    Reads the **soil table and **position_data table from the input position CSV file.
    Creates a stratigraphy plot with:
    - Colored background layers (from plot_Color column)
    - Step plot showing constant qc within each layer
    - Only layers below seabed (depth >= 0)
    - Linear scale x-axis

    Args:
        position: Position name (e.g., 'A01')
        input_file_path: Path to input_Position_*.csv file (independent from driveability CSV)
        output_dir: Directory to save the plot (optional)

    Returns:
        None (creates and saves plot, or returns early on error)
    """
    print(f"\n{'='*60}")
    print(f"Plotting Soil Profile for Position {position}")
    print(f"{'='*60}")

    # Parse the input CSV file to get the **soil table
    try:
        tables = parse_results_csv(input_file_path)
    except Exception as e:
        print(f"Error parsing input file {input_file_path}: {e}")
        return

    # Get the position_data table to read zSeabed
    position_data_table = tables.get('position_data')
    zSeabed = None
    if position_data_table is not None and 'zSeabed' in position_data_table.columns:
        try:
            zSeabed_value = position_data_table['zSeabed'].iloc[0]
            zSeabed = pd.to_numeric(zSeabed_value, errors='coerce')
            if pd.notna(zSeabed):
                print(f"Found zSeabed from position_data: {zSeabed} m")
            else:
                print(f"Warning: zSeabed value is NaN, will use first layer elevation instead")
                zSeabed = None
        except (IndexError, KeyError) as e:
            print(f"Warning: Could not read zSeabed from position_data: {e}")
            zSeabed = None
    else:
        print(f"Warning: No position_data table or zSeabed column found")

    # Get the soil table
    soil_table = tables.get('soil')
    if soil_table is None:
        print(f"Warning: No **soil table found in {input_file_path}")
        return

    print(f"Found soil table with {len(soil_table)} layers")

    # Debug: Show available columns
    print(f"Available columns: {list(soil_table.columns)}")

    # Debug: Show first few rows
    print(f"First 3 rows of soil table:")
    print(soil_table[['z_top', 'qc', 'plot_Color']].head(3))

    # Check required columns
    required_cols = ['z_top', 'qc', 'plot_Color']
    missing_cols = [col for col in required_cols if col not in soil_table.columns]
    if missing_cols:
        print(f"Error: Missing required columns in soil table: {missing_cols}")
        return

    # Extract data and convert to numeric
    print(f"\nExtracting and converting soil data...")
    z_top_lat = pd.to_numeric(soil_table['z_top'], errors='coerce')  # Elevation in LAT (meters)
    qc_pa = pd.to_numeric(soil_table['qc'], errors='coerce')  # Cone resistance in Pa

    # Handle plot_Color: convert to string and replace NaN/invalid values with default color
    colors = soil_table['plot_Color'].copy()
    # Replace NaN with a default color before converting to string
    colors = colors.fillna('#d3d3d3')  # Default gray color
    colors = colors.astype(str)
    # Also handle any remaining 'nan' strings or empty strings
    colors = colors.replace(['nan', 'NaN', 'NA', ''], '#d3d3d3')

    # Extract geoUnit column for labeling (if available)
    geo_units = None
    if 'geoUnit' in soil_table.columns:
        geo_units = soil_table['geoUnit'].copy()
        geo_units = geo_units.fillna('')  # Replace NaN with empty string
        geo_units = geo_units.astype(str)
        print(f"Found geoUnit column with {geo_units.notna().sum()} entries")
    else:
        print(f"Warning: No geoUnit column found in soil table")

    # Debug: Check data types and sample values
    print(f"  z_top: {z_top_lat.dtype}, non-null: {z_top_lat.notna().sum()}/{len(z_top_lat)}")
    print(f"  qc: {qc_pa.dtype}, non-null: {qc_pa.notna().sum()}/{len(qc_pa)}")
    if len(z_top_lat) > 0:
        print(f"  Sample z_top[0]: {z_top_lat.iloc[0]}")
        print(f"  Sample qc[0]: {qc_pa.iloc[0]}")

    # Convert depth: subtract zSeabed elevation to get depth below seabed
    if len(z_top_lat) == 0:
        print("Error: No soil layers found")
        return

    # Filter out rows where z_top or qc are NaN before processing
    valid_initial = z_top_lat.notna() & qc_pa.notna()
    print(f"  Valid rows (both z_top and qc present): {valid_initial.sum()}/{len(valid_initial)}")

    z_top_lat = z_top_lat[valid_initial]
    qc_pa = qc_pa[valid_initial]
    colors = colors[valid_initial]
    if geo_units is not None:
        geo_units = geo_units[valid_initial]

    if len(z_top_lat) == 0:
        print("Error: No valid z_top or qc data found")
        return

    # Use zSeabed from position_data if available, otherwise use first layer elevation
    if zSeabed is not None:
        seabed_elevation = zSeabed
        print(f"  Using zSeabed from position_data: {seabed_elevation} m")
    else:
        seabed_elevation = z_top_lat.iloc[0]
        print(f"  Using first layer elevation as seabed: {seabed_elevation} m")

    depth_below_seabed = seabed_elevation - z_top_lat  # Positive depth below seabed

    # Convert qc from Pa to MPa
    qc_mpa = qc_pa / 1e6

    # Filter to only include layers below seabed (positive depth) and valid qc
    valid_mask = (depth_below_seabed >= 0) & (qc_mpa > 0)
    print(f"  Layers below seabed with valid qc: {valid_mask.sum()}/{len(valid_mask)}")

    depth_below_seabed = depth_below_seabed[valid_mask]
    qc_mpa = qc_mpa[valid_mask]
    colors = colors[valid_mask]
    if geo_units is not None:
        geo_units = geo_units[valid_mask]

    if len(depth_below_seabed) == 0:
        print("Error: No valid soil data to plot below seabed")
        return

    print(f"Plotting {len(depth_below_seabed)} soil layers")
    print(f"Depth range: {depth_below_seabed.min():.2f} to {depth_below_seabed.max():.2f} m")
    print(f"qc range: {qc_mpa.min():.2f} to {qc_mpa.max():.2f} MPa")

    # Get the max qc value to set plot range
    max_qc = qc_mpa.max()

    # Create figure
    fig = go.Figure()

    # Build step plot data with constant qc within each layer
    # Each layer has same qc at top and bottom
    qc_step = []
    depth_step = []

    for i in range(len(depth_below_seabed) - 1):
        depth_start = depth_below_seabed.iloc[i]
        depth_end = depth_below_seabed.iloc[i + 1]
        qc_value = qc_mpa.iloc[i]

        # Add points for constant qc within layer (horizontal line)
        qc_step.extend([qc_value, qc_value])
        depth_step.extend([depth_start, depth_end])

    # Add last layer (extends to end)
    if len(depth_below_seabed) > 0:
        qc_step.append(qc_mpa.iloc[-1])
        depth_step.append(depth_below_seabed.iloc[-1])

    # Plot each layer as a separate trace with its own background color spanning full width
    for i in range(len(depth_below_seabed) - 1):
        # Get depth range for this layer (from current to next layer)
        depth_start = depth_below_seabed.iloc[i]
        depth_end = depth_below_seabed.iloc[i + 1]
        layer_color = str(colors.iloc[i])  # Ensure it's a string

        # Validate color (ensure it starts with # for hex colors)
        if not layer_color.startswith('#') and not layer_color.startswith('rgb'):
            layer_color = '#d3d3d3'  # Default gray

        print(f"  Layer {i}: depth {depth_start:.2f}-{depth_end:.2f} m, qc: {qc_mpa.iloc[i]:.2f} MPa, color: {layer_color}")

        # Add a filled rectangle spanning full width for this layer's background
        fig.add_trace(go.Scatter(
            x=[0, max_qc * 1.1, max_qc * 1.1, 0, 0],  # Full width from 0 to max qc + 10%
            y=[depth_start, depth_start, depth_end, depth_end, depth_start],
            fill='toself',
            fillcolor=layer_color,
            mode='none',
            showlegend=False,
            hoverinfo='skip'
        ))

    # Add the main qc profile line on top (step plot with constant qc per layer)
    fig.add_trace(go.Scatter(
        x=qc_step,
        y=depth_step,
        mode='lines',
        name='qc Profile',
        line=dict(color='black', width=2),
        hovertemplate='<b>Depth: %{y:.2f} m bsf</b><br>qc: %{x:.2f} MPa<br><extra></extra>'
    ))

    # Add geoUnit labels for contiguous zones (if geoUnit column exists)
    if geo_units is not None and len(geo_units) > 0:
        print(f"\nIdentifying contiguous geoUnit zones for labeling...")

        # Reset index for easier iteration
        geo_units_reset = geo_units.reset_index(drop=True)
        depth_reset = depth_below_seabed.reset_index(drop=True)

        # Identify contiguous zones where the same geoUnit appears
        zones = []  # List of (geoUnit_name, start_depth, end_depth)
        current_unit = None
        zone_start = None

        for i in range(len(geo_units_reset)):
            unit = geo_units_reset.iloc[i]
            depth = depth_reset.iloc[i]

            # Skip empty or invalid geoUnit names
            if not unit or unit in ['', 'nan', 'NaN', 'NA']:
                # Close current zone if any
                if current_unit is not None:
                    zones.append((current_unit, zone_start, depth))
                    current_unit = None
                    zone_start = None
                continue

            # Start new zone or continue current zone
            if unit != current_unit:
                # Close previous zone if any
                if current_unit is not None:
                    zones.append((current_unit, zone_start, depth))

                # Start new zone
                current_unit = unit
                zone_start = depth

        # Close the last zone (extends to the bottom of the last layer)
        if current_unit is not None and len(depth_reset) > 0:
            # Get the bottom of the last layer
            last_depth = depth_reset.iloc[-1]
            zones.append((current_unit, zone_start, last_depth))

        print(f"  Found {len(zones)} contiguous geoUnit zones")

        # Add text annotations for each zone (at top of zone, right side, no box)
        for unit_name, start_depth, end_depth in zones:
            print(f"    Zone: {unit_name} from {start_depth:.2f} to {end_depth:.2f} m (label at top: {start_depth:.2f} m)")

            # Add text annotation on the right side of the plot, at the top of each zone
            fig.add_annotation(
                x=0.95,  # 95% from left edge (right side)
                y=start_depth,  # Top of the zone
                xref='paper',  # x position relative to plot width (0 to 1)
                yref='y',  # y position in data coordinates (depth)
                text=f'<b>{unit_name}</b>',
                showarrow=False,
                font=dict(size=12, color='black', family='Arial'),
                xanchor='right',  # Anchor text to the right
                yanchor='top'  # Anchor text to the top
            )

    # Update layout - LINEAR scale, not logarithmic
    fig.update_layout(
        title=f'Soil Profile - Position {position}',
        xaxis=dict(
            title='Cone Resistance qc [MPa]',
            gridcolor='lightgray',
            showgrid=True,
            linecolor='#5a5a5a',
            linewidth=2,
            mirror=True,
            range=[0, max_qc * 1.1]  # Linear scale from 0 to max + 10%
        ),
        yaxis=dict(
            title='Depth below seabed [m]',
            autorange='reversed',  # Depth increases downward
            gridcolor='lightgray',
            showgrid=True,
            linecolor='#5a5a5a',
            linewidth=2,
            mirror=True
        ),
        hovermode='closest',
        template='plotly_white',
        width=900,
        height=1000,
        plot_bgcolor='white',
        paper_bgcolor='white'
    )

    # Save plot if output directory is specified
    if output_dir:
        output_path = output_dir / f'soil_profile_{position}.html'
        fig.write_html(str(output_path))
        print(f"Soil profile plot saved to: {output_path}")

    # Show interactive plot
    fig.show()
    print(f"✓ Soil profile plot displayed")


def plot_driveability_results(tables: Dict[str, pd.DataFrame], position: str,
                      selected_methods: List[str], selected_bounds: List[str],
                      output_dir: Path = None,
                      monopile_weights: dict = None,
                      position_info: dict = None,
                      soil_profile_data: dict = None,
                      gripper_data: dict = None,
                      abandonment_data: dict = None,
                      refusal_data: dict = None,
                      show_plot: bool = True):
    """
    Create comprehensive interactive driveability analysis plot for a single position.

    Generates a 2x4 subplot grid with synchronized y-axes for easy comparison:

    ROW 1 (Driving Parameters):
    - Col 1: Static Resistance to Driving (SRD) [MN] vs Depth with soil profile overlay
    - Col 2: Blowcount Rate [bl/25cm] vs Depth with hard driving & refusal thresholds
    - Col 3: Input Energy [kJ/blow] vs Depth
    - Col 4: Position Information Table (weights, depths, thresholds)

    ROW 2 (Cumulative Metrics & Assessment):
    - Col 1: Total Energy [GJ] vs Depth (cumulative energy consumption)
    - Col 2: Cumulative Blows vs Depth (total hammer blows)
    - Col 3: SRD [kN] vs Depth (logarithmic scale) with weight thresholds
    - Col 4: Self-Weight Penetration & Pile Run Assessment Table

    Key Features:
    - Color coding: Each SRD method gets unique color (MD=blue, MY=red, AH=green)
    - Line styles: Soil bounds differentiated (be=solid, lb=dash, ub=dot)
    - Soil overlay: qc profile with color-coded layers and geoUnit labels on SRD plot
    - Horizontal lines: Target depth, gripper release, MP abandonment, refusal risks
    - Weight thresholds: MP+ILT, MP+Hammer, MP only shown on log SRD plot
    - A3 landscape format: Optimized for printing/reporting (1587x1123 px)

    Args:
        tables: Dictionary of DataFrames parsed from driveability results CSV
        position: Position name (e.g., 'A01', 'A02')
        selected_methods: List of SRD methods to plot (e.g., ['MD', 'AH'])
        selected_bounds: List of soil bounds to plot (e.g., ['lb', 'be', 'ub'])
        output_dir: Directory to save HTML plot (optional, uses PLOTS_OUTPUT_DIR if None)
        monopile_weights: Dict mapping position names to monopile weights in tonnes
        position_info: Dict with hammer details, target depth, blowcount rate
        soil_profile_data: Dict with qc profile, layers, colors, geoUnits (optional)
                          Keys: depth_below_seabed, qc_mpa, colors, qc_step,
                                depth_step, max_qc, geo_unit_zones
        gripper_data: Dict mapping positions to gripper release depths (optional)
        abandonment_data: Dict mapping positions to MP abandonment depths (optional)
        refusal_data: Dict mapping positions to refusal risk depths by duration (optional)
                     Format: {position: {'1hr': depth, '24hr': depth, ...}}
        show_plot: If True, opens plot in browser; if False, only saves to file (default: True)

    Output:
        - Saves interactive HTML plot as: Installation_Driveability_{position}.html
        - Opens plot in default browser for immediate review
        - Console output shows plot statistics and file location

    Notes:
        - All subplots share synchronized y-axes for consistent zooming
        - Uses banker's rounding avoidance (ROUND_HALF_UP) for displayed values
        - Gracefully handles missing optional data (skips overlay if unavailable)
    """
    # ===================================================================================================
    # INITIALIZATION AND SETUP
    # ===================================================================================================

    # Extract target depth from summary table
    target_depth = None
    summary_table = tables.get('results_PileDrivingAnalysis_Summary')
    if summary_table is not None and 'targetdepth' in summary_table.columns:
        try:
            target_depth = pd.to_numeric(summary_table['targetdepth'].iloc[0])
            print(f"Target depth: {target_depth} m")
        except (ValueError, IndexError):
            print("Warning: Could not extract target depth from summary table")

    # Define color scheme for SRD methods (consistent across all plots)
    method_colors = {
        'MD': '#1f77b4',  # blue
        'MY': '#d62728',  # red
        'AH': '#2ca02c'   # green
    }

    # Define line style patterns for soil bounds (Plotly format)
    bound_dashes = {
        'be': 'solid',      # Best estimate: solid line (most prominent)
        'lb': 'dash',       # Lower bound: dashed line
        'ub': 'dot'         # Upper bound: dotted line
    }

    # ===================================================================================================
    # CREATE SUBPLOT GRID (2 rows × 4 columns)
    # ===================================================================================================
    # Layout: Row 1 - SRD, Blowcount, Energy, Info Table
    #         Row 2 - Total Energy, Cumulative Blows, SRD (log), Assessment Table
    # A3 landscape format (420mm × 297mm) at 96 DPI = 1587px × 1123px
    # Synchronized y-axes enable consistent zooming across all plots
    fig = make_subplots(
        rows=2, cols=4,
        shared_yaxes='all',  # Critical: enables synchronized zooming/panning
        row_heights=[0.5, 0.5],  # Equal height distribution
        column_widths=[0.23, 0.23, 0.23, 0.31],  # Plots: 23% each, Tables: 31%
        horizontal_spacing=0.06,
        vertical_spacing=0.12,
        subplot_titles=(None, None, None, None, None, None, None, None),
        specs=[[{'type': 'xy'}, {'type': 'xy'}, {'type': 'xy'}, {'type': 'table'}],
               [{'type': 'xy'}, {'type': 'xy'}, {'type': 'xy'}, {'type': 'table'}]]
    )
    plot_count = 0

    # Prepare soil profile data for later overlay (added after traces to determine axis ranges)
    soil_layer_data = None
    if soil_profile_data is not None:
        print("Preparing soil profile overlay for SRD plot...")
        soil_layer_data = {
            'depth_below_seabed': soil_profile_data['depth_below_seabed'],
            'colors': soil_profile_data['colors'],
            'qc_step': soil_profile_data['qc_step'],
            'depth_step': soil_profile_data['depth_step'],
            'max_qc': soil_profile_data['max_qc'],
            'geo_unit_zones': soil_profile_data['geo_unit_zones']
        }

    # ===================================================================================================
    # EXTRACT AND PLOT DRIVEABILITY DATA FOR SELECTED METHODS/BOUNDS
    # ===================================================================================================
    # Storage for later assessment calculations
    methods_to_plot = []
    bounds_to_plot = []
    ruts_to_plot = []
    depths_to_plot = []

    # Iterate through all tables and plot selected method/bound combinations
    for table_name, df in tables.items():
        method, bound = extract_method_and_bound(table_name)

        # Skip if not in user's selection
        if method not in selected_methods or bound not in selected_bounds:
            continue

        # Validate required columns exist in this table
        if 'Depth' not in df.columns or 'Rut' not in df.columns or 'Blowcount_rate' not in df.columns or 'Input_Energy' not in df.columns:
            continue

        # Convert all data columns to numeric, handling errors gracefully
        depth = pd.to_numeric(df['Depth'], errors='coerce')
        rut = pd.to_numeric(df['Rut'], errors='coerce')  # Static Resistance to Driving in MN
        blowcount_rate = pd.to_numeric(df['Blowcount_rate'], errors='coerce') / 4.0  # Convert bl/m to bl/25cm
        blowcount_rate_per_m = pd.to_numeric(df['Blowcount_rate'], errors='coerce')  # Keep bl/m for cumulative calculation
        input_energy = pd.to_numeric(df['Input_Energy'], errors='coerce')  # Energy per blow in kJ
        # Use Cumulative_input_energy column if present, else fallback to cumsum of Input_Energy
        cumulative_input_energy = pd.to_numeric(df['Cumulative_input_Energy'], errors='coerce') / 1_000_000

        # Remove NaN values
        mask = ~(depth.isna() | rut.isna() | blowcount_rate.isna() | blowcount_rate_per_m.isna() | input_energy.isna() | cumulative_input_energy.isna())
        depth = depth[mask]
        rut = rut[mask]
        blowcount_rate = blowcount_rate[mask]
        blowcount_rate_per_m = blowcount_rate_per_m[mask]
        input_energy = input_energy[mask]
        cumulative_input_energy = cumulative_input_energy[mask]

        if len(depth) == 0:
            continue

        # Calculate cumulative blows
        # For each depth interval, calculate blows = blowcount_rate * depth_increment
        # Then accumulate
        depth_increments = depth.diff().fillna(depth.iloc[0] if len(depth) > 0 else 0)
        blows_per_interval = blowcount_rate_per_m * depth_increments
        cumulative_blows = blows_per_interval.cumsum()

        color = method_colors.get(method, '#000000')
        dash = bound_dashes.get(bound, 'solid')
        label = f'{method}_{bound}'

        # Plot in row 1, column 1: SRD [MN] vs Depth [mbsb]
        fig.add_trace(
            go.Scatter(
                x=rut,
                y=depth,
                mode='lines',
                name=label,
                line=dict(color=color, width=2, dash=dash),
                legendgroup=label,
                hovertemplate='<b>%{fullData.name}</b><br>SRD: %{x:.2f} MN<br>Depth: %{y:.2f} m<br><extra></extra>'
            ),
            row=1, col=1
        )
        # Plot in row 1, column 2: Blowcount Rate [bl/25cm] vs Depth [mbsb]
        fig.add_trace(
            go.Scatter(
                x=blowcount_rate,
                y=depth,
                mode='lines',
                name=label,
                line=dict(color=color, width=2, dash=dash),
                legendgroup=label,
                showlegend=False,
                hovertemplate='<b>%{fullData.name}</b><br>Blowcount Rate: %{x:.2f} bl/25cm<br>Depth: %{y:.2f} m<br><extra></extra>'
            ),
            row=1, col=2
        )
        # Plot in row 1, column 3: Input_Energy [kJ/blow] vs Depth [mbsb]
        fig.add_trace(
            go.Scatter(
                x=input_energy,
                y=depth,
                mode='lines',
                name=label,
                line=dict(color=color, width=2, dash=dash),
                legendgroup=label,
                showlegend=False,
                hovertemplate='<b>%{fullData.name}</b><br>Input Energy: %{x:.2f} kJ/blow<br>Depth: %{y:.2f} m<br><extra></extra>'
            ),
            row=1, col=3
        )
        # New plot: Total Energy [GJ] vs Depth [mbsb] in row 2, col 1
        fig.add_trace(
            go.Scatter(
                x=cumulative_input_energy,
                y=depth,
                mode='lines',
                name=label,
                line=dict(color=color, width=2, dash=dash),
                legendgroup=label,
                showlegend=False,
                hovertemplate='<b>%{fullData.name}</b><br>Total Energy: %{x:.2f} GJ<br>Depth: %{y:.2f} m<br><extra></extra>'
            ),
            row=2, col=1
        )
        # New plot: Cumulative Blows vs Depth [mbsb] in row 2, col 2
        fig.add_trace(
            go.Scatter(
                x=cumulative_blows,
                y=depth,
                mode='lines',
                name=label,
                line=dict(color=color, width=2, dash=dash),
                legendgroup=label,
                showlegend=False,
                hovertemplate='<b>%{fullData.name}</b><br>Cumulative Blows: %{x:.0f}<br>Depth: %{y:.2f} m<br><extra></extra>'
            ),
            row=2, col=2
        )
        # New plot: SRD [kN] vs Depth [mbsb] in row 2, col 3 (logarithmic scale)
        rut_kn = rut * 1000  # Convert MN to kN
        fig.add_trace(
            go.Scatter(
                x=rut_kn,
                y=depth,
                mode='lines',
                name=label,
                line=dict(color=color, width=2, dash=dash),
                legendgroup=label,
                showlegend=False,
                hovertemplate='<b>%{fullData.name}</b><br>SRD: %{x:.0f} kN<br>Depth: %{y:.2f} m<br><extra></extra>'
            ),
            row=2, col=3
        )
        # Collect for intersection calculation
        methods_to_plot.append(method)
        bounds_to_plot.append(bound)
        ruts_to_plot.append(rut)
        depths_to_plot.append(depth)
        plot_count += 1

    if plot_count == 0:
        print("Warning: No data found to plot!")
        return

    # Add target depth horizontal line if available (to all subplots)
    if target_depth is not None:
        fig.add_hline(
            y=target_depth,
            line_dash="dash",
            line_color="black",
            line_width=2,
            annotation_text=f"Target Depth: {target_depth} m",
            annotation_position="top left",
            annotation=dict(font_size=11, font_color="black"),
            row=1, col=1
        )
        fig.add_hline(
            y=target_depth,
            line_dash="dash",
            line_color="black",
            line_width=2,
            showlegend=False,
            row=1, col=2
        )
        fig.add_hline(
            y=target_depth,
            line_dash="dash",
            line_color="black",
            line_width=2,
            showlegend=False,
            row=1, col=3
        )
        fig.add_hline(
            y=target_depth,
            line_dash="dash",
            line_color="black",
            line_width=2,
            showlegend=False,
            row=2, col=1
        )
        fig.add_hline(
            y=target_depth,
            line_dash="dash",
            line_color="black",
            line_width=2,
            showlegend=False,
            row=2, col=2
        )
        fig.add_hline(
            y=target_depth,
            line_dash="dash",
            line_color="black",
            line_width=2,
            showlegend=False,
            row=2, col=3
        )

    # Use hardcoded values instead of user input
    hard_driving_val = HARD_DRIVING_BLOWCOUNT
    refusal_val = REFUSAL_BLOWCOUNT

    # Add vertical lines to second subplot if values are provided
    if hard_driving_val is not None:
        fig.add_vline(
            x=hard_driving_val,
            line_dash="dash",
            line_color="orange",
            line_width=3,
            annotation_text="hard driving",
            annotation_position="top left",
            annotation=dict(font_size=12, font_color="orange", textangle=-90),
            row=1, col=2
        )
    if refusal_val is not None:
        fig.add_vline(
            x=refusal_val,
            line_dash="dash",
            line_color="red",
            line_width=3,
            annotation_text="refusal",
            annotation_position="top left",
            annotation=dict(font_size=12, font_color="red", textangle=-90),
            row=1, col=2
        )

    # Get monopile weight
    mp_weight = None
    if monopile_weights is not None:
        mp_weight = monopile_weights.get(position, None)

    # === CALCULATE SWP AND PILE RUN ASSESSMENT ===
    # Call the calculation function to get all assessment results
    assessment = calculate_swp_and_pile_run_assessment(
        methods_to_plot=methods_to_plot,
        bounds_to_plot=bounds_to_plot,
        ruts_to_plot=ruts_to_plot,
        depths_to_plot=depths_to_plot,
        mp_weight=mp_weight
    )

    # Extract results for easier reference
    swp_mp_ilt_depths = assessment['swp_mp_ilt_depths']
    pile_run_at_hammer_placement = assessment['pile_run_at_hammer_placement']
    swp_mp_hammer_depths = assessment['swp_mp_hammer_depths']
    pile_run_risk_top = assessment['pile_run_risk_top']
    pile_run_risk_bottom = assessment['pile_run_risk_bottom']
    mp_lift_tool_total_weight_kn = assessment['mp_lift_tool_total_weight_kn']
    mp_hammer_total_weight_kn = assessment['mp_hammer_total_weight_kn']

    # === ADD VERTICAL REFERENCE LINES TO PLOT ===
    # Add vertical lines to row 2, col 3 (SRD in kN log scale) for weight references
    if mp_weight is not None:
        # Add line for MP + additional weight + lifting tool weight (dashed black line)
        fig.add_vline(
            x=mp_lift_tool_total_weight_kn,
            line_dash="dash",
            line_color="black",
            line_width=2,
            annotation_text="MP+add.+lift tool",
            annotation_position="top right",
            annotation=dict(font_size=11, font_color="black", textangle=-90),
            row=2, col=3
        )

        # Add line for MP + additional weight + hammer (solid black line)
        fig.add_vline(
            x=mp_hammer_total_weight_kn,
            line_dash="solid",
            line_color="black",
            line_width=2,
            annotation_text="MP+add.+hammer",
            annotation_position="top right",
            annotation=dict(font_size=11, font_color="black", textangle=-90),
            row=2, col=3
        )

        # Add invisible traces for legend in row 2, col 3 only
        # These create legend entries without showing actual data points
        fig.add_trace(
            go.Scatter(
                x=[None],
                y=[None],
                mode='lines',
                line=dict(color='black', width=2, dash='dash'),
                name='MP+ILT',
                showlegend=True,
                legendgroup='srd_lines',
                legend='legend2',  # Use separate legend
                hoverinfo='skip'
            ),
            row=2, col=3
        )

        fig.add_trace(
            go.Scatter(
                x=[None],
                y=[None],
                mode='lines',
                line=dict(color='black', width=2, dash='solid'),
                name='MP+Hammer',
                showlegend=True,
                legendgroup='srd_lines',
                legend='legend2',  # Use separate legend
                hoverinfo='skip'
            ),
            row=2, col=3
        )

    # === CREATE ASSESSMENT RESULTS TABLE ===
    weight_table = go.Table(
        columnwidth=[0.526, 0.158, 0.158, 0.158],  # Control column widths: 50% for first column, 17% each for LB/BE/UB
        header=dict(
            values=['<b>Evaluated  Depth [mbsb]</b>', '<b>LB</b>', '<b>BE</b>', '<b>UB</b>'],
            fill_color=['#b3c6e7', '#e6f2ff', '#e6f2ff', '#e6f2ff'],  # Nicer color for first row
            align='center',
            font=dict(size=13, color='black', family='Arial')
        ),
        cells=dict(
            values=[
                ['<b>SWP MP + ILT</b>', '<b>Pile run @ hammer placement</b>', '<b>SWP MP + Hammer</b>', '<b>Pile run risk top</b>', '<b>Pile run risk bottom</b>'],
                [swp_mp_ilt_depths['LB'], pile_run_at_hammer_placement['LB'], swp_mp_hammer_depths['LB'], pile_run_risk_top['LB'], pile_run_risk_bottom['LB']],  # LB column
                [swp_mp_ilt_depths['BE'], pile_run_at_hammer_placement['BE'], swp_mp_hammer_depths['BE'], pile_run_risk_top['BE'], pile_run_risk_bottom['BE']],  # BE column
                [swp_mp_ilt_depths['UB'], pile_run_at_hammer_placement['UB'], swp_mp_hammer_depths['UB'], pile_run_risk_top['UB'], pile_run_risk_bottom['UB']]   # UB column
            ],
            fill_color='white',
            align=['left', 'center', 'center'],
            font=dict(size=11.9, color='black', family='Arial'),
            height=30
        )
    )

    fig.add_trace(weight_table, row=2, col=4)
    fig.update_xaxes(visible=False, row=1, col=4)
    fig.update_yaxes(visible=False, row=1, col=4)
    fig.update_xaxes(visible=False, row=2, col=4)
    fig.update_yaxes(visible=False, row=2, col=4)
    # ===============================================================================

    # --- Info panel table in row 1, col 4 ---
    # Get gripper penetration value for this position (if available)
    # Use decimal module to avoid banker's rounding (round half up instead of round half to even)
    gripper_penetration = ''
    if gripper_data and position in gripper_data:
        # Round using ROUND_HALF_UP to avoid banker's rounding (e.g., 16.95 → 17.0)
        rounded_val = float(Decimal(str(gripper_data[position])).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP))
        gripper_penetration = f"{rounded_val:.1f}"

    # Get MP abandonment penetration value for this position (if available)
    # Use decimal module to avoid banker's rounding
    abandonment_penetration = ''
    if abandonment_data and position in abandonment_data:
        # Round using ROUND_HALF_UP to avoid banker's rounding
        rounded_val = float(Decimal(str(abandonment_data[position])).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP))
        abandonment_penetration = f"{rounded_val:.1f}"

    # Get refusal risk depths for this position (if available)
    refusal_1hr = ''
    refusal_24hr = ''
    refusal_48hr = ''
    refusal_7days = ''
    if refusal_data and position in refusal_data:
        pos_refusal = refusal_data[position]
        if '1hr' in pos_refusal:
            val = pos_refusal['1hr']
            if val == "No risk":
                refusal_1hr = "No risk"
            else:
                rounded_val = float(Decimal(str(val)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP))
                refusal_1hr = f"{rounded_val:.1f}"
        if '24hr' in pos_refusal:
            val = pos_refusal['24hr']
            if val == "No risk":
                refusal_24hr = "No risk"
            else:
                rounded_val = float(Decimal(str(val)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP))
                refusal_24hr = f"{rounded_val:.1f}"
        if '48hr' in pos_refusal:
            val = pos_refusal['48hr']
            if val == "No risk":
                refusal_48hr = "No risk"
            else:
                rounded_val = float(Decimal(str(val)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP))
                refusal_48hr = f"{rounded_val:.1f}"
        if '7days' in pos_refusal:
            val = pos_refusal['7days']
            if val == "No risk":
                refusal_7days = "No risk"
            else:
                rounded_val = float(Decimal(str(val)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP))
                refusal_7days = f"{rounded_val:.1f}"

    # ===================================================================================================
    # CREATE INFORMATION AND ASSESSMENT TABLES (Col 4)
    # ===================================================================================================

    # --- POSITION INFORMATION TABLE (Row 1, Col 4) ---
    # Contains project-specific data, installation parameters, and risk thresholds
    info_table = go.Table(
        columnwidth=[0.7, 0.3],
        header=dict(
            values=['<b>Info</b>', '<b>Value</b>'],
            fill_color='#b3c6e7',
            align='center',
            font=dict(size=13, color='black', family='Arial')
        ),
        cells=dict(
            values=[
                ['Position', 'Monopile Weight [t]', 'Target Penetration Depth [m]', 'Hammer', 'Hammer Weight [t]', 'Target Blowcount Rate [bl/25cm]', 'Min. Penetration for Gripper release [m]', 'Min. Penetration for MP abandonment [m]', '1hr instal. pause - Refusal risk depth [m]', '24hr instal. pause - Refusal risk depth [m]', '48hr instal. pause - Refusal risk depth [m]', '7days instal. pause - Refusal risk depth [m]'],
                [
                    position,
                    f"{mp_weight:.1f}" if mp_weight is not None else '',
                    f"{position_info.get('target_penetration_depth', '')}",
                    position_info.get('hammer_name', ''),
                    f"{position_info.get('hammer_weight', '')}",
                    f"{position_info.get('target_blowcount_rate', '')}",
                    gripper_penetration,
                    abandonment_penetration,
                    refusal_1hr,
                    refusal_24hr,
                    refusal_48hr,
                    refusal_7days
                ]
            ],
            # Use different background colors: white for general info and refusal risk, light blue for minimum penetration rows
            # Provide colors for each column: [column1_colors, column2_colors]
            fill_color=[
                ['white', 'white', 'white', 'white', 'white', 'white', '#e6f2ff', '#e6f2ff', 'white', 'white', 'white', 'white'],  # Info column
                ['white', 'white', 'white', 'white', 'white', 'white', '#e6f2ff', '#e6f2ff', 'white', 'white', 'white', 'white']   # Value column
            ],
            align=['left', 'center'],
            font=dict(size=12, color='black', family='Arial'),
            height=30
        )
    )
    fig.add_trace(info_table, row=1, col=4)

    # Update layout for all subplots
    fig.update_layout(
        title={
            'text': f'Position {position}',
            'x': 0.5,
            'xanchor': 'center',
            'font': {'size': 18, 'family': 'Arial, sans-serif'}
        },
        xaxis=dict(
            title='SRD [MN]',
            gridcolor='lightgray',
            showgrid=True,
            linecolor='#5a5a5a',
            linewidth=2,
            mirror=True
        ),
        xaxis2=dict(
            title='Blowcount Rate [bl/25cm]',
            gridcolor='lightgray',
            showgrid=True,
            linecolor='#5a5a5a',
            linewidth=2,
            mirror=True,
            range=[0, 250]
        ),
        xaxis3=dict(
            title='Input Energy [kJ/blow]',
            gridcolor='lightgray',
            showgrid=True,
            linecolor='#5a5a5a',
            linewidth=2,
            mirror=True
        ),
        xaxis4=dict(
            title='Total Energy [GJ]',
            gridcolor=None,
            showgrid=False,
            linecolor='#5a5a5a',
            linewidth=2,
            mirror=True
        ),
        yaxis=dict(
            title='Depth [mbsb]',
            autorange='reversed',
            rangemode='tozero',
            gridcolor='lightgray',
            showgrid=True,
            dtick=5,
            linecolor='#5a5a5a',
            linewidth=2,
            mirror=True,
            showticklabels=True
        ),
        yaxis2=dict(
            title='Depth [mbsb]',
            autorange='reversed',
            rangemode='tozero',
            gridcolor='lightgray',
            showgrid=True,
            dtick=5,
            linecolor='#5a5a5a',
            linewidth=2,
            mirror=True,
            showticklabels=True
        ),
        yaxis3=dict(
            title='Depth [mbsb]',
            autorange='reversed',
            rangemode='tozero',
            gridcolor='lightgray',
            showgrid=True,
            dtick=5,
            linecolor='#5a5a5a',
            linewidth=2,
            mirror=True,
            showticklabels=True
        ),
        yaxis4=dict(
            title='Depth [mbsb]',
            autorange='reversed',
            rangemode='tozero',
            gridcolor=None,
            showgrid=False,
            dtick=5,
            linecolor='#5a5a5a',
            linewidth=2,
            mirror=True,
            showticklabels=True
        ),
        hovermode='closest',
        template='plotly_white',
        legend=dict(
            yanchor="bottom",
            y=0.02,
            xanchor="right",
            x=0.98,
            bgcolor="rgba(255, 255, 255, 0.9)",
            bordercolor="Black",
            borderwidth=1
        ),
        legend2=dict(
            yanchor="bottom",
            y=0.04,  # Position within row 2 area (0.5 to 1.0 for row 2)
            xanchor="left",
            x=0.51,  # Position within col 3 area (approximately 0.69-0.92 for col 3)
            bgcolor="rgba(255, 255, 255, 0.9)",
            bordercolor="Black",
            borderwidth=1,
            orientation="v",  # Vertical orientation
            font=dict(size=10),
            title=dict(text="", font=dict(size=10))  # No title for cleaner look
        ),
        width=1587,   # A3 landscape width (420mm) at 96 DPI
        height=1123,  # A3 landscape height (297mm) at 96 DPI
        font=dict(size=11),  # Slightly smaller font for better fit on A3
        plot_bgcolor='white',
        paper_bgcolor='white'
    )

    # Update layout for row 2, col 1 (Total Energy plot)
    fig.update_xaxes(
        title='Total Energy [GJ]',
        gridcolor='lightgray',
        showgrid=True,
        linecolor='#5a5a5a',
        linewidth=2,
        mirror=True,
        showticklabels=True,
        row=2, col=1
    )
    fig.update_yaxes(
        title='Depth [mbsb]',
        autorange='reversed',
        rangemode='tozero',
        gridcolor='lightgray',
        showgrid=True,
        dtick=5,
        linecolor='#5a5a5a',
        linewidth=2,
        mirror=True,
        showticklabels=True,
        row=2, col=1
    )

    # Update layout for row 2, col 2 (Cumulative Blows plot)
    fig.update_xaxes(
        title='Cumulative Blows',
        gridcolor='lightgray',
        showgrid=True,
        linecolor='#5a5a5a',
        linewidth=2,
        mirror=True,
        showticklabels=True,
        row=2, col=2
    )
    fig.update_yaxes(
        title='Depth [mbsb]',
        autorange='reversed',
        rangemode='tozero',
        gridcolor='lightgray',
        showgrid=True,
        dtick=5,
        linecolor='#5a5a5a',
        linewidth=2,
        mirror=True,
        showticklabels=True,
        row=2, col=2
    )

    # Update layout for row 2, col 3 (SRD in kN with log scale)
    fig.update_xaxes(
        title='SRD [kN]',
        type='log',
        range=[4, 5.176],  # log10(10000) = 4, log10(150000) ≈ 5.176
        tickmode='array',
        tickvals=[10000, 20000, 30000, 40000, 50000, 60000, 70000, 80000, 90000, 100000, 150000],
        ticktext=['10k', '', '', '', '', '', '', '', '', '100k', '150k'],
        gridcolor='lightgray',
        showgrid=True,
        linecolor='#5a5a5a',
        linewidth=2,
        mirror=True,
        showticklabels=True,
        row=2, col=3
    )
    fig.update_yaxes(
        title='Depth [mbsb]',
        autorange='reversed',
        rangemode='tozero',
        gridcolor='lightgray',
        showgrid=True,
        dtick=5,
        linecolor='#5a5a5a',
        linewidth=2,
        mirror=True,
        showticklabels=True,
        row=2, col=3
    )


    fig.add_annotation(
        text="<b>Self weight penetration and pile run assessment</b>",
        xref="paper",
        yref="paper",
        x=0.75,
        y=0.44,
        showarrow=False,
        font=dict(size=14, color='black', family='Arial'),
        xanchor='left',
        yanchor='bottom'
    )

    # ===================================================================================================
    # SOIL PROFILE OVERLAY ON SRD SUBPLOT (Row 1, Col 1)
    # ===================================================================================================
    # If soil profile data is available, overlay it on the SRD plot to show geological context.
    # This includes:
    # 1. Color-coded soil layers as background rectangles
    # 2. qc (cone resistance) profile as grey line on secondary x-axis
    # 3. geoUnit labels positioned at layer boundaries

    if soil_layer_data is not None:
        print("Finalizing soil profile overlay on SRD plot...")

        # Extract prepared soil data
        depth_below_seabed = soil_layer_data['depth_below_seabed']
        colors = soil_layer_data['colors']
        qc_step = soil_layer_data['qc_step']
        depth_step = soil_layer_data['depth_step']
        max_qc = soil_layer_data['max_qc']
        geo_unit_zones = soil_layer_data['geo_unit_zones']

        # --- DETERMINE DEPTH RANGE FOR SOIL OVERLAY ---
        # Match the depth range of SRD data traces to avoid unnecessary soil layer display
        # Find maximum depth from all plotted SRD traces
        max_depth = 0
        for trace in fig.data:
            # Check if this trace has depth data (y values) and is not a table trace
            if hasattr(trace, 'y') and trace.y is not None and len(trace.y) > 0:
                try:
                    trace_max = max(trace.y)
                    if trace_max > max_depth:
                        max_depth = trace_max
                except:
                    pass  # Skip traces without valid y data

        # Ensure soil profile extends at least to target depth (full installation depth)
        if target_depth:
            if max_depth < target_depth:
                max_depth = target_depth
        elif max_depth == 0:
            # Fallback if no data and no target depth
            max_depth = 40

        print(f"  Using max depth: {max_depth:.2f} m for soil profile overlay")

        # --- FILTER SOIL DATA TO MATCH DEPTH RANGE ---
        # Only display soil layers within the SRD plot depth range (optimization + clarity)
        depth_below_seabed_filtered = depth_below_seabed[depth_below_seabed <= max_depth]

        # Ensure the filtered depths extend to max_depth for complete layer coverage
        if len(depth_below_seabed_filtered) > 0 and depth_below_seabed_filtered.iloc[-1] < max_depth:
            # Add max_depth as a final boundary point
            depth_below_seabed_filtered = pd.concat([
                depth_below_seabed_filtered,
                pd.Series([max_depth])
            ]).reset_index(drop=True)

        # Filter qc step plot data to match depth range
        qc_step_filtered = []
        depth_step_filtered = []
        for i, d in enumerate(depth_step):
            if d <= max_depth:
                qc_step_filtered.append(qc_step[i])
                depth_step_filtered.append(d)

        # Extend qc profile to max_depth if necessary (maintains constant value in deepest layer)
        if len(depth_step_filtered) > 0 and depth_step_filtered[-1] < max_depth:
            # Extend the last qc value down to max_depth
            qc_step_filtered.append(qc_step_filtered[-1])
            depth_step_filtered.append(max_depth)

        # --- ADD SOIL LAYER BACKGROUNDS TO SRD SUBPLOT ---
        # Color-coded rectangles provide geological context behind SRD curves
        # Applied ONLY to row 1, col 1 to avoid cluttering other subplots
        num_layers = len(depth_below_seabed_filtered) - 1

        # Iterate through soil layers and add as background shapes
        for i in range(num_layers):
            if i >= len(colors):
                break  # Safety check for color array bounds

            # Define layer depth boundaries
            depth_start = depth_below_seabed_filtered.iloc[i]
            depth_end = depth_below_seabed_filtered.iloc[i + 1] if i + 1 < len(depth_below_seabed_filtered) else max_depth
            layer_color = str(colors.iloc[i])

            # Validate color format (must be hex or rgb)
            if not layer_color.startswith('#') and not layer_color.startswith('rgb'):
                layer_color = '#d3d3d3'  # Default gray for invalid colors

            # Add rectangle shape spanning full subplot width (x domain: 0 to 1)
            fig.add_shape(
                type="rect",
                xref="x domain",  # Use domain coordinates (0 to 1) for x
                yref="y",  # Use data coordinates for y (depth)
                x0=0,  # Left edge of subplot
                x1=1,  # Right edge of subplot
                y0=depth_start,
                y1=depth_end,
                fillcolor=layer_color,
                line=dict(width=0),
                layer="below",  # Draw below traces
                row=1, col=1
            )

        # --- EXTEND DEEPEST SOIL LAYER TO PLOT BOTTOM ---
        # Handle case where filtered depths extend beyond available color data
        # This ensures no white gap at the bottom of the plot
        if len(depth_below_seabed_filtered) > 0 and len(colors) > 0:
            # Find the last layer that has a corresponding color
            last_colored_index = min(len(colors) - 1, len(depth_below_seabed_filtered) - 2)
            if last_colored_index >= 0 and last_colored_index < len(depth_below_seabed_filtered) - 1:
                last_depth = depth_below_seabed_filtered.iloc[last_colored_index + 1]
                last_color = str(colors.iloc[last_colored_index])

                # Validate color format
                if not last_color.startswith('#') and not last_color.startswith('rgb'):
                    last_color = '#d3d3d3'

                # Extend the deepest layer's color to the bottom of the plot
                if last_depth < max_depth:
                    fig.add_shape(
                        type="rect",
                        xref="x domain",
                        yref="y",
                        x0=0,
                        x1=1,
                        y0=last_depth,
                        y1=max_depth,  # Extend to bottom of subplot
                        fillcolor=last_color,
                        line=dict(width=0),
                        layer="below",
                        row=1, col=1
                    )

        # --- ADD QC PROFILE AS SECONDARY X-AXIS TRACE ---
        # Grey line showing cone resistance (qc) variation with depth
        # Uses xaxis7 (secondary x-axis at top) while sharing yaxis with SRD plot
        # Step plot style maintains constant qc within each soil layer
        fig.add_trace(
            go.Scatter(
                x=qc_step_filtered,
                y=depth_step_filtered,
                mode='lines',
                name='qc Profile',
                line=dict(color='grey', width=2, dash='solid'),
                xaxis='x7',  # Secondary x-axis (configured below)
                yaxis='y',   # Share y-axis with SRD plot (row 1, col 1)
                hovertemplate='<b>qc Profile</b><br>qc: %{x:.2f} MPa<br>Depth: %{y:.2f} m<br><extra></extra>',
                showlegend=False  # Don't show in legend
            )
        )

        # --- CONFIGURE SECONDARY X-AXIS FOR QC PROFILE ---
        # Create xaxis7 overlaying the primary x-axis (SRD) on row 1, col 1
        # Positioned at top of subplot to avoid interference with SRD axis at bottom
        # Grey styling distinguishes it from primary SRD axis
        fig.update_layout(
            xaxis7=dict(
                title=dict(text='qc [MPa]', font=dict(color='grey', size=11)),
                overlaying='x',  # Overlay on primary x-axis (SRD axis)
                side='top',      # Position at top of subplot
                anchor='y',      # Anchor to primary y-axis (depth)
                range=[0, max_qc * 1.1],  # Extend range by 10% for visual clearance
                showline=True,   # Display axis line for clear separation
                linecolor='grey',
                linewidth=2,
                showgrid=False,  # Disable grid to avoid cluttering SRD plot
                zeroline=False,
                showticklabels=True,
                tickmode='linear',
                tick0=0,         # Start ticks at 0 MPa
                dtick=20,        # Major tick marks every 20 MPa
                ticks='outside',
                ticklen=5,
                tickwidth=2,
                tickcolor='grey',
                tickfont=dict(color='grey', size=10),
                mirror=False,
                # Minor ticks for finer scale reading
                minor=dict(
                    tickmode='linear',
                    tick0=0,
                    dtick=5,     # Minor tick marks every 5 MPa
                    ticks='outside',
                    ticklen=3,   # Shorter than major ticks
                    tickwidth=1,
                    tickcolor='grey',
                    showgrid=False
                )
            )
        )

        # Explicitly lock y-axis range to prevent auto-extension beyond max_depth
        # Ensures consistent vertical scale across all synchronized subplots
        fig.update_yaxes(range=[max_depth, 0], row=1, col=1)

        # --- ADD GEOUNIT LABELS TO SRD SUBPLOT ---
        # Label geological unit zones at their top boundaries (right-aligned)
        # Only label zones visible within the current depth range
        for unit_name, start_depth, end_depth in geo_unit_zones:
            if start_depth <= max_depth:
                # Position label at 95% of subplot width (right side, avoiding overlap with data)
                # x domain coordinates: 0 = left edge, 1 = right edge of subplot
                fig.add_annotation(
                    x=0.95,          # 95% from left edge of subplot
                    y=start_depth,   # At the top boundary of the geological unit
                    xref='x domain', # Use normalized subplot coordinates (0-1)
                    yref='y',        # Use actual depth data coordinates
                    text=f'<b>{unit_name}</b>',
                    showarrow=False,
                    font=dict(size=11, color='black', family='Arial'),
                    xanchor='right', # Right-align text to prevent extending beyond subplot edge
                    yanchor='top'    # Top-align text to position at layer boundary
                )

        print(f"  ✓ Added qc profile, {num_layers} soil layers, and {len([z for z in geo_unit_zones if z[1] <= max_depth])} geoUnit labels to SRD subplot")

    # ===================================================================================================
    # SAVE AND DISPLAY PLOT
    # ===================================================================================================
    # Save interactive HTML plot with position-specific filename
    if output_dir:
        print(f"  [4/4] Saving plot to file...")
        output_path = output_dir / f'Installation_Driveability_{position}.html'
        fig.write_html(str(output_path))
        print(f"        ✓ Saved: {output_path}")

    # Open interactive plot in default browser for immediate review (optional)
    if show_plot:
        print(f"  [4/4] Opening plot in browser...")
        fig.show()
        print(f"        ✓ Displayed with {plot_count} traces per subplot")
    else:
        print(f"  [4/4] Plot generation completed ({plot_count} traces per subplot)")


def generate_pile_run_summary(
    selected_positions: List[str],
    position_tables: Dict[str, Dict[str, pd.DataFrame]],
    selected_methods: List[str],
    selected_bounds: List[str],
    monopile_weights: dict,
    output_dir: Path
) -> None:
    """
    Generate a summary CSV/Excel file with pile run assessment results for all requested positions.

    Creates a file with three header rows showing:
    - Row 1: Assessment categories (SWP MP + ILT, Pile run @hammer placement, etc.)
    - Row 2: Method_Bound combinations (lb_MY, be_MY, ub_MY, etc.)
    - Row 3: Units (m for all depth columns)

    The summary includes 5 assessment categories, each with values for all bound combinations:
    1. SWP MP + ILT: Self-weight penetration with internal lifting tool
    2. Pile run @hammer placement: Risk assessment at hammer placement depth
    3. SWP MP + Hammer: Self-weight penetration with hammer weight
    4. Pile run risk top: First depth where pile run risk initiates
    5. Pile run risk bottom: Depth where pile run risk zone ends

    Args:
        selected_positions: List of position names to include (e.g., ['A01', 'A02'])
        position_tables: Dictionary mapping position names to their parsed CSV tables
        selected_methods: List of SRD methods selected (e.g., ['MD', 'MY', 'AH'])
        selected_bounds: List of soil bounds selected (e.g., ['lb', 'be', 'ub'])
        monopile_weights: Dictionary mapping position names to monopile weights
        output_dir: Directory where the summary file will be saved

    Output File:
        - Filename: summary_pile_run.xlsx (Excel format for proper multi-row headers)
        - Location: output_dir / summary_pile_run.xlsx
    """
    print("\n" + "="*60)
    print("GENERATING PILE RUN SUMMARY FILE")
    print("="*60)

    # Determine all method-bound combinations from selected methods and bounds
    method_order = sorted(selected_methods)

    # Define bound order explicitly as lb, be, ub (not alphabetical)
    # Filter to only include bounds that are selected
    desired_bound_order = ['lb', 'be', 'ub']
    bound_order = [b for b in desired_bound_order if b in selected_bounds]

    # Create header rows - NEW LAYOUT: Group by assessment category, then bounds
    header_row_1 = ['Position']  # Category level
    header_row_2 = ['']  # Bound level
    header_row_3 = ['text']  # Units level

    # Define assessment categories (5 categories, each with all bounds)
    assessment_categories = [
        'SWP MP + ILT',
        'Pile run @hammer placement',
        'SWP MP + Hammer',
        'Pile run risk top',
        'Pile run risk bottom'
    ]

    # Create method label for header row 2
    # If multiple methods selected, show all methods to indicate conservative selection
    # If single method, just show that method
    if len(selected_methods) == 1:
        method_label = selected_methods[0]
    else:
        # Show all methods to indicate conservative value across all
        method_label = '+'.join(sorted(selected_methods))

    # NEW LAYOUT: For each assessment category, add all bounds
    # This groups assessments together vertically for easier reading
    for category in assessment_categories:
        for bound in bound_order:
            header_row_1.append(category)
            # Format: lb_MD or lb_MD+MY+AH (showing conservative across methods)
            header_row_2.append(f'{bound}_{method_label}')
            header_row_3.append('m')  # All values are in meters

    # Initialize data storage
    data_rows = []

    # Process each position
    print(f"  Total positions to process: {len(selected_positions)}")
    for position in selected_positions:
        print(f"  Processing {position}...")

        # Get tables for this position
        tables = position_tables.get(position, {})
        if not tables:
            print(f"    Warning: No data found for {position}")
            # Calculate correct number of empty columns: 5 categories × number of bounds
            row_data = [position] + [''] * (len(assessment_categories) * len(bound_order))
            data_rows.append(row_data)
            print(f"    Added empty row for {position}. Total rows now: {len(data_rows)}")
            continue

        # Get monopile weight
        mp_weight = monopile_weights.get(position)
        if mp_weight is None:
            print(f"    Warning: No monopile weight found for {position}")
            # Still add row with empty values
            row_data = [position] + [''] * (len(assessment_categories) * len(bound_order))
            data_rows.append(row_data)
            print(f"    Added empty row for {position} (no weight). Total rows now: {len(data_rows)}")
            continue

        # Prepare data for assessment calculation
        methods_to_plot = []
        bounds_to_plot = []
        ruts_to_plot = []
        depths_to_plot = []

        # Extract data from tables for selected methods and bounds
        for table_name, df in tables.items():
            method, bound = extract_method_and_bound(table_name)

            if method not in selected_methods or bound not in selected_bounds:
                continue

            # Validate required columns
            if 'Depth' not in df.columns or 'Rut' not in df.columns:
                continue

            # Convert to numeric
            depth = pd.to_numeric(df['Depth'], errors='coerce')
            rut = pd.to_numeric(df['Rut'], errors='coerce')

            # Remove NaN values
            mask = ~(depth.isna() | rut.isna())
            depth = depth[mask]
            rut = rut[mask]

            if len(depth) == 0:
                continue

            methods_to_plot.append(method)
            bounds_to_plot.append(bound)
            ruts_to_plot.append(rut)
            depths_to_plot.append(depth)

        # Calculate pile run assessments
        if methods_to_plot:
            assessment = calculate_swp_and_pile_run_assessment(
                methods_to_plot=methods_to_plot,
                bounds_to_plot=bounds_to_plot,
                ruts_to_plot=ruts_to_plot,
                depths_to_plot=depths_to_plot,
                mp_weight=mp_weight
            )

            # Build row data - NEW LAYOUT: Group by assessment category
            row_data = [position]

            # Helper function to convert string values to numbers where appropriate
            def convert_to_number(value):
                """Convert string numbers to float, keep text as-is"""
                if isinstance(value, str):
                    try:
                        # Try to convert to float if it's a numeric string
                        return float(value)
                    except (ValueError, AttributeError):
                        # Keep as text if conversion fails (e.g., "No risk", "Yes", etc.)
                        return value
                return value

            # For each assessment category, add all bounds
            # This matches the new header layout

            # 1. SWP MP + ILT - all bounds
            for bound in bound_order:
                bound_key = bound.upper()
                value = assessment['swp_mp_ilt_depths'].get(bound_key, '')
                row_data.append(convert_to_number(value))

            # 2. Pile run @hammer placement - all bounds
            for bound in bound_order:
                bound_key = bound.upper()
                value = assessment['pile_run_at_hammer_placement'].get(bound_key, '')
                row_data.append(convert_to_number(value))

            # 3. SWP MP + Hammer - all bounds
            for bound in bound_order:
                bound_key = bound.upper()
                value = assessment['swp_mp_hammer_depths'].get(bound_key, '')
                row_data.append(convert_to_number(value))

            # 4. Pile run risk top - all bounds
            for bound in bound_order:
                bound_key = bound.upper()
                value = assessment['pile_run_risk_top'].get(bound_key, '')
                row_data.append(convert_to_number(value))

            # 5. Pile run risk bottom - all bounds
            for bound in bound_order:
                bound_key = bound.upper()
                value = assessment['pile_run_risk_bottom'].get(bound_key, '')
                row_data.append(convert_to_number(value))

            data_rows.append(row_data)
            print(f"    Successfully added row for {position}. Total rows now: {len(data_rows)}")
        else:
            print(f"    Warning: No valid data to calculate assessments for {position}")
            row_data = [position] + [''] * (len(assessment_categories) * len(bound_order))
            data_rows.append(row_data)
            print(f"    Added empty row for {position}. Total rows now: {len(data_rows)}")

    # Create DataFrame
    print(f"\n  Creating DataFrame with {len(data_rows)} rows...")
    df_summary = pd.DataFrame(data_rows, columns=header_row_2)
    print(f"  DataFrame shape: {df_summary.shape} (rows × columns)")
    print(f"  Position column values: {df_summary.iloc[:, 0].tolist()}")

    # Save to Excel with multi-row headers
    output_file = output_dir / 'summary_pile_run.xlsx'
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write the dataframe starting from row 4 (0-indexed: row 3) - after 3 header rows
            df_summary.to_excel(writer, sheet_name='Pile Run Summary', startrow=3, index=False, header=False)

            # Get the worksheet to add custom headers
            worksheet = writer.sheets['Pile Run Summary']

            # Write header row 1 (categories)
            for col_idx, value in enumerate(header_row_1, start=1):
                worksheet.cell(row=1, column=col_idx, value=value)

            # Write header row 2 (method_bound)
            for col_idx, value in enumerate(header_row_2, start=1):
                worksheet.cell(row=2, column=col_idx, value=value)

            # Write header row 3 (units)
            for col_idx, value in enumerate(header_row_3, start=1):
                worksheet.cell(row=3, column=col_idx, value=value)

            # Format headers (bold)
            from openpyxl.styles import Font, Alignment
            for row in [1, 2, 3]:
                for col in range(1, len(header_row_1) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        print(f"\n✓ Summary file saved: {output_file}")
        print(f"  Positions included: {len(data_rows)}")
        print(f"  Columns: {len(header_row_1)}")

    except Exception as e:
        print(f"\n✗ Error saving summary file: {e}")
        # Fallback: Save as CSV (won't have perfect multi-row headers but will work)
        csv_file = output_dir / 'summary_pile_run.csv'
        try:
            # Combine headers into single rows with appropriate separators
            combined_header = []
            for i in range(len(header_row_2)):
                combined_header.append(f"{header_row_1[i]}|{header_row_2[i]}|{header_row_3[i]}")

            df_csv = pd.DataFrame(data_rows, columns=combined_header)
            df_csv.to_csv(csv_file, index=False)
            print(f"\n✓ Fallback: Summary file saved as CSV: {csv_file}")
        except Exception as e2:
            print(f"\n✗ Error saving CSV fallback: {e2}")


def main():
    """
    Main execution function orchestrating the entire driveability analysis workflow.

    EXECUTION FLOW:
    ===============

    1. INITIALIZATION
       - Set root directory from MONOPILE_ROOT_DIR constant
       - Fallback to local directory if network path unavailable
       - Scan for position folders (A01, A02, etc.)

    2. POSITION SELECTION
       - Display available positions to user
       - Prompt for position selection (comma-separated or 'all')
       - Validate and store selected positions

    3. DATA LOADING (Efficient - only for selected positions)
       - Load gripper penetration data (gripper release + MP abandonment depths)
       - Parse driveability CSV files for each selected position
       - Extract available SRD methods (MD, MY, AH) and soil bounds (lb, be, ub)
       - Store target penetration depths for refusal risk calculations

    4. METHOD & BOUND SELECTION
       - Display available SRD methods across all selected positions
       - Prompt for method selection (comma-separated or 'all')
       - Display available soil bounds across all selected positions
       - Prompt for bound selection (comma-separated or 'all')

    5. REFUSAL RISK DATA LOADING
       - Load refusal risk assessment data from Excel file
       - Read sheets: '1 hr', '24 hr', '48 hr', '7 days'
       - Apply conservative logic across selected methods
       - Store refusal depths for each position and pause duration

    6. POSITION PROCESSING LOOP
       For each selected position:
       a) Extract soil profile data from input_Position_{Position}.csv
          - qc profile, layer colors, geoUnit classifications

       b) Call plot_driveability_results() to create comprehensive plot:
          - 2x4 subplot grid with synchronized y-axes
          - SRD plots with soil profile overlay
          - Blowcount rates with thresholds
          - Energy consumption metrics
          - Assessment tables with SWP and pile run evaluations

       c) Save plot as: Installation_Driveability_{Position}.html

       d) Display interactive plot in browser

    ERROR HANDLING:
    ===============
    - Network path unavailable: Falls back to local directory
    - No positions found: Exits with warning message
    - Missing data files: Graceful degradation, continues without optional overlays
    - Parse errors: Logs warnings but continues processing other positions

    USER INTERACTION:
    =================
    - Position selection: User chooses which positions to analyze
    - Method selection: User chooses SRD calculation methods to compare
    - Bound selection: User chooses soil parameter bounds to include
    - All selections support 'all' keyword for full analysis

    CONSOLE OUTPUT:
    ===============
    - Progress messages for each major step
    - Data loading confirmations with counts
    - Warning messages for missing/invalid data
    - Plot save locations and success confirmations
    - Visual separators (=== lines) for readability
    """
    # Start timing for performance tracking
    start_time = time.time()

    # Clear parse cache from any previous runs
    clear_parse_cache()

    # Set the root directory
    root_dir = MONOPILE_ROOT_DIR

    # Fallback to local directory if network path is unavailable
    if not root_dir.exists():
        print(f"Warning: Network directory not found: {root_dir}")
        print("Using local HOW03 directory instead...")
        root_dir = Path(__file__).parent

    # Get position folders
    positions = get_position_folders(root_dir)

    if not positions:
        print("No position folders found!")
        return

    print(f"Found {len(positions)} positions: {[p[0] for p in positions]}")

    # Ask for position selection FIRST (before parsing any files)
    print("\n" + "="*60)
    print("SELECT POSITION(S) TO PLOT")
    print("="*60)
    print(f"\nAvailable positions: {', '.join([p[0] for p in positions])}")
    selected_positions_input = input("Enter positions to plot (comma-separated, e.g., A01,A02 or 'all'): ").strip()

    if selected_positions_input.lower() == 'all':
        selected_positions = [p[0] for p in positions]
    else:
        selected_positions = [p.strip().upper() for p in selected_positions_input.split(',') if p.strip()]

    print(f"\n✓ Selected {len(selected_positions)} position(s) for analysis")

    # ===================================================================================================
    # LOAD POSITION-SPECIFIC DATA (OPTIMIZED - ONLY SELECTED POSITIONS)
    # ===================================================================================================
    load_start = time.time()

    # Load penetration data ONLY for selected positions (efficient!)
    # Returns two dictionaries: gripper release and MP abandonment
    gripper_data, abandonment_data = get_gripper_penetration_for_positions(selected_positions)

    # Now only parse CSV files for SELECTED positions (much faster!)
    print(f"\nParsing data for {len(selected_positions)} selected position(s)...")
    all_methods = set()
    all_bounds = set()
    position_tables = {}
    target_depths = {}  # Store target depths for refusal risk calculation

    for p_name, p_path in positions:
        if p_name not in selected_positions:
            continue  # Skip positions not selected
        csv_file = p_path / f"results_PileDrivingAnalysis-{p_name}.csv"
        if csv_file.exists():
            print(f"  Reading {p_name}...")
            tables = parse_results_csv(csv_file)
            position_tables[p_name] = tables
            methods, bounds = get_available_methods_and_bounds(tables)
            all_methods.update(methods)
            all_bounds.update(bounds)

            # Extract target depth for this position (needed for refusal risk calculation)
            summary_table = tables.get('results_PileDrivingAnalysis_Summary')
            if summary_table is not None and 'targetdepth' in summary_table.columns:
                try:
                    target_depths[p_name] = float(pd.to_numeric(summary_table['targetdepth'].iloc[0]))
                except (ValueError, IndexError):
                    pass

    methods = sorted(list(all_methods))
    bounds = sorted(list(all_bounds))

    parse_time = time.time() - load_start
    print(f"\n✓ Data loading completed in {parse_time:.2f} seconds")


    # User selection for methods and bounds
    print("\n" + "="*60)
    print("SELECT SRD METHODS AND SOIL BOUNDS")
    print("="*60)
    print(f"\nAvailable SRD methods: {', '.join(methods)}")
    selected_methods_input = input("Enter methods to plot (comma-separated, e.g., MD,AH or 'all'): ").strip()

    if selected_methods_input.lower() == 'all':
        selected_methods = methods
    else:
        selected_methods = [m.strip() for m in selected_methods_input.split(',') if m.strip()]

    print(f"Selected methods: {selected_methods}")

    print(f"\nAvailable soil bounds: {', '.join(bounds)}")
    selected_bounds_input = input("Enter bounds to plot (comma-separated, e.g., lb,ub or 'all'): ").strip()

    if selected_bounds_input.lower() == 'all':
        selected_bounds = bounds
    else:
        selected_bounds = [b.strip().lower() for b in selected_bounds_input.split(',') if b.strip()]

    print(f"Selected bounds: {selected_bounds}")

    # Ask user if they want to open plots in browser (can slow down execution)
    print(f"\nOpen plots in browser? (Opening {len(selected_positions)} plot(s) can be slow)")
    show_plots_input = input("Open in browser? (y/n, default=n): ").strip().lower()
    show_plots = show_plots_input in ['y', 'yes']
    if not show_plots:
        print("✓ Plots will be saved to file only (faster execution)")

    # Load refusal risk data for selected positions and methods (efficient!)
    # This reads from the Excel file sheets: '1 hr', '24 hr', '48 hr', '7 days'
    refusal_data = get_refusal_risk_for_positions(selected_positions, selected_methods, target_depths)

    # ===================================================================================================
    # LOAD COMMON DATA ONCE (PERFORMANCE OPTIMIZATION)
    # ===================================================================================================
    # Load monopile weights ONCE for all selected positions instead of once per position
    print(f"\nLoading monopile weights for {len(selected_positions)} position(s)...")
    monopile_weights = get_monopile_weights(MONOPILE_WEIGHTS_FILE, selected_positions)

    # ===================================================================================================
    # PLOTTING LOOP - GENERATE INTERACTIVE PLOTS FOR EACH POSITION
    # ===================================================================================================
    plot_start = time.time()
    print(f"\n{'='*60}")
    print(f"GENERATING PLOTS FOR {len(selected_positions)} POSITION(S)")
    print(f"{'='*60}")

    # --- MAIN LOOP: Process each selected position ---
    for idx, position in enumerate(selected_positions, 1):
        position_start = time.time()
        print(f"\n[{idx}/{len(selected_positions)}] Processing position: {position}")
        print(f"{'='*60}")

        # Get corresponding tables
        tables = position_tables.get(position, {})

        # Extract soil profile data from input position file (for overlay on SRD plot)
        soil_profile_data = None
        try:
            position_dir = MONOPILE_ROOT_DIR / position
            input_files = list(position_dir.glob(f'input_Position_*_{position}.csv'))
            if input_files:
                input_file = input_files[0]
                print(f"  [1/4] Loading soil profile data...")
                soil_profile_data = extract_soil_profile_data(input_file)
                if soil_profile_data:
                    print(f"        ✓ Extracted {len(soil_profile_data['depth_below_seabed'])} soil layers")
        except Exception as e:
            print(f"  Warning: Could not extract soil profile data: {e}")

        # Get position info once (reused in plot function)
        print(f"  [2/4] Extracting position information...")
        position_info = get_position_info(tables, position, selected_methods, selected_bounds)

        # Plot Rut vs Depth with soil profile overlay
        print(f"  [3/4] Generating plot traces...")
        plot_driveability_results(
            tables=tables,
            position=position,
            selected_methods=selected_methods,
            selected_bounds=selected_bounds,
            output_dir=PLOTS_OUTPUT_DIR,
            monopile_weights=monopile_weights,
            position_info=position_info,
            soil_profile_data=soil_profile_data,
            gripper_data=gripper_data,
            abandonment_data=abandonment_data,
            refusal_data=refusal_data,
            show_plot=show_plots
        )

        position_time = time.time() - position_start
        print(f"\n✓ Position {position} completed in {position_time:.2f} seconds")

    # ===================================================================================================
    # GENERATE PILE RUN SUMMARY FILE
    # ===================================================================================================
    # Create summary Excel/CSV file with pile run assessment results for all positions
    try:
        generate_pile_run_summary(
            selected_positions=selected_positions,
            position_tables=position_tables,
            selected_methods=selected_methods,
            selected_bounds=selected_bounds,
            monopile_weights=monopile_weights,
            output_dir=PLOTS_OUTPUT_DIR
        )
    except Exception as e:
        print(f"\n✗ Error generating pile run summary: {e}")
        import traceback
        traceback.print_exc()

    # ===================================================================================================
    # EXECUTION SUMMARY
    # ===================================================================================================
    total_time = time.time() - start_time
    plot_time = time.time() - plot_start

    print(f"\n{'='*60}")
    print(f"EXECUTION SUMMARY")
    print(f"{'='*60}")
    print(f"Data loading time:    {parse_time:.2f} seconds")
    print(f"Plotting time:        {plot_time:.2f} seconds")
    print(f"Total execution time: {total_time:.2f} seconds")
    print(f"Average per position: {plot_time/len(selected_positions):.2f} seconds")
    print(f"{'='*60}")
    print(f"✓ Successfully generated {len(selected_positions)} plot(s)")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()


# ===================================================================================================
# CODE STRUCTURE QUICK REFERENCE
# ===================================================================================================
"""
FUNCTION HIERARCHY AND CALL ORDER:
-----------------------------------

main()
├── get_position_folders() → Scan for position directories
├── User Input: Select positions
├── get_gripper_penetration_for_positions() → Load gripper/abandonment data
├── FOR each selected position:
│   ├── parse_results_csv() → Extract driveability tables
│   ├── get_available_methods_and_bounds() → Identify methods/bounds
│   └── Extract target depths
├── User Input: Select methods and bounds
├── get_refusal_risk_for_positions() → Load refusal risk data
└── FOR each selected position:
    ├── extract_soil_profile_data() → Read soil layers
    │   ├── parse_results_csv() → Parse input position file
    │   └── Process soil profile (qc, colors, geoUnits)
    ├── get_monopile_weights() → Read MP weights from Excel
    ├── get_position_info() → Extract hammer and target details
    └── plot_driveability_results() → Create interactive plot
        ├── Create 2x4 subplot grid
        ├── Plot SRD traces for all method/bound combinations
        ├── Add target depth lines
        ├── Add hard driving and refusal thresholds
        ├── calculate_swp_and_pile_run_assessment() → Calculate assessments
        │   ├── Calculate SWP MP + ILT
        │   ├── Calculate pile run at hammer placement
        │   ├── Calculate SWP MP + Hammer
        │   ├── Calculate pile run risk top/bottom
        │   └── Apply conservative logic across methods
        ├── Add weight threshold reference lines
        ├── Create assessment results table
        ├── Create position information table
        ├── Add gripper/abandonment horizontal lines
        ├── Add refusal risk horizontal lines
        ├── Overlay soil profile on SRD subplot
        │   ├── Filter soil data to depth range
        │   ├── Add color-coded layer backgrounds
        │   ├── Add qc profile trace on secondary x-axis
        │   ├── Configure secondary x-axis (xaxis7)
        │   └── Add geoUnit labels
        ├── Update plot layout and axes
        ├── Save as Installation_Driveability_{Position}.html
        └── Display interactive plot in browser

HELPER FUNCTIONS:
-----------------
- parse_results_csv() → Parse multi-table CSV files
- extract_method_and_bound() → Extract method/bound from table name
- get_available_methods_and_bounds() → Identify available methods/bounds
- get_position_folders() → Scan for position directories
- get_monopile_weights() → Read MP weights from Excel
- get_position_info() → Extract hammer and target information
- get_gripper_penetration_for_positions() → Read penetration thresholds
- get_refusal_risk_for_positions() → Read refusal risk depths
- extract_soil_profile_data() → Process soil profile for overlay
- calculate_swp_and_pile_run_assessment() → Calculate installation assessments
- plot_soil_profile() → Independent soil profile plotting (not used in main flow)
- plot_driveability_results() → Main plotting function

DATA FLOW:
----------
Input Files → Parse → Extract Methods/Bounds → User Selection → 
Calculate Assessments → Create Plots → Save HTML → Display

KEY ALGORITHMS:
---------------
1. Conservative Value Selection (calculate_swp_and_pile_run_assessment):
   - SWP depths: Select DEEPEST across methods (most conservative penetration)
   - Pile run risks: Select SHALLOWEST for top (earliest risk occurrence)
   - Pile run risks: Select DEEPEST for bottom (longest risk zone)

2. Soil Profile Overlay (plot_driveability_results):
   - Filter soil layers to match SRD plot depth range
   - Create step plot for constant qc within layers
   - Add color-coded rectangles as background shapes
   - Overlay qc profile on secondary x-axis (top of subplot)
   - Add geoUnit labels at layer boundaries

3. Refusal Risk Assessment (get_refusal_risk_for_positions):
   - Read multiple Excel sheets (1hr, 24hr, 48hr, 7days)
   - Find position row and method columns
   - Extract difference from target depth (negative values)
   - Calculate absolute refusal depth: target_depth + difference
   - Apply conservative logic: most negative difference across methods

OUTPUT FILES:
-------------
- Installation_Driveability_{Position}.html for each position
- Interactive Plotly plots with 2x4 subplot grid
- A3 landscape format (1587px × 1123px)
- Synchronized y-axes for consistent zooming

CUSTOMIZATION POINTS:
---------------------
1. Lines 108-109: HARD_DRIVING_BLOWCOUNT, REFUSAL_BLOWCOUNT
2. Lines 113-115: INTERNAL_LIFTING_TOOL, HAMMER_WEIGHT, ADDITIONAL_WEIGHT
3. Lines 119-125: File paths (weights, root dir, output dir)
4. Lines 128-135: Gripper and refusal risk data sources
5. Lines 1556-1561: Method colors and line dash patterns
6. Lines 1572-1582: Subplot grid configuration
"""
